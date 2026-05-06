"""
Serviço de importação em lote de cards via planilha Excel.
Suporta criação de Clientes, Pessoas e Cards vinculados em uma única operação.
"""
import io
from datetime import datetime
from unicodedata import normalize as unicode_normalize

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.card import Card
from app.models.client import Client
from app.models.person import Person
from app.models.user import User
from app.models.card_list_history import CardListHistory
from app.schemas.card import CardImportRowResult, CardImportResponse

TARGET_LIST_ID = 22
TARGET_BOARD_ID = 6
EXEMPLO_PREFIX = "[EXEMPLO]"

# Seções e suas cores de cabeçalho
SECTION_COLORS = {
    "card":   "1E3A5F",  # azul escuro
    "client": "14532D",  # verde escuro
    "person": "3B0764",  # roxo escuro
    "owner":  "78350F",  # âmbar escuro
}

# Colunas: (header, exemplo, obrigatório, seção, campo_person_model)
# campo_person_model: usado para mapear colunas de telefone na criação da Person
TEMPLATE_COLUMNS = [
    # ── Card ──────────────────────────────────────────────────────────────────
    ("Título do Card",      "Proposta Empresa XYZ",               True,  "card",   None),
    ("Descrição",           "Interesse no produto A",              False, "card",   None),
    ("Tipo de Negócio",     "Nova Venda",                          True,  "card",   None),
    ("Canal de Aquisição",  "Indicacao",                           True,  "card",   None),
    ("Detalhe do Canal",    "Indicação do cliente XYZ",            True,  "card",   None),
    # ── Empresa ───────────────────────────────────────────────────────────────
    ("Nome da Empresa",     "Empresa XYZ LTDA",                    True,  "client", None),
    ("CNPJ/CPF",            "12.345.678/0001-90",                  False, "client", None),
    ("Setor / Segmento",    "Tecnologia",                          True,  "client", None),
    ("Website",             "www.empresa.com.br",                  False, "client", None),
    ("Cidade",              "São Paulo",                           False, "client", None),
    ("Estado (UF)",         "SP",                                  False, "client", None),
    ("Nº de Funcionários",  "Ate 50 colaboradores",                False, "client", None),
    ("Receita Anual",       "Ate R$ 10 milhoes",                   False, "client", None),
    ("CNAE",                "6201-5/01",                           False, "client", None),
    # ── Contato ───────────────────────────────────────────────────────────────
    ("Nome do Contato",     "Carlos Oliveira",                     True,  "person", None),
    ("Cargo",               "Diretor Comercial",                   False, "person", None),
    ("E-mail Comercial",    "carlos@empresa.com.br",               True,  "person", "email_commercial"),
    ("Telefone Principal",  "(11) 88888-8888",                     False, "person", "phone"),
    ("WhatsApp",            "(11) 99999-9999",                     False, "person", "phone_whatsapp"),
    ("Telefone Comercial",  "(11) 3333-4444",                      False, "person", "phone_commercial"),
    ("Telefone Alternativo","(11) 3333-5555",                      False, "person", "phone_alternative"),
    ("Telefone Extra 1",    "(11) 3333-6666",                      False, "person", "phone_extra1"),
    ("Telefone Extra 2",    "(11) 3333-7777",                      False, "person", "phone_extra2"),
    ("LinkedIn",            "linkedin.com/in/carlos",              False, "person", "linkedin"),
    # ── Responsáveis ──────────────────────────────────────────────────────────
    ("Vendedor (nome exato)", "",                                  False, "owner",  None),
    ("SDR (nome exato)",      "",                                  False, "owner",  None),
]

VALID_CHANNELS = {
    "inbound": "Inbound",
    "outbound": "Outbound",
    "indicacao": "Indicacao",
    "indicação": "Indicacao",
    "parcerias": "Parcerias",
    "eventos": "Eventos",
    "base": "Base",
}

VALID_DEAL_TYPES = {
    "nova venda": "Nova Venda",
    "cross sell": "Cross Sell",
    "up sell": "Up Sell",
}

VALID_EMPLOYEE_COUNTS = [
    "Ate 50 colaboradores",
    "51-100 colaboradores",
    "101-300 colaboradores",
    "301-600 colaboradores",
    "601-1.000 colaboradores",
    "Acima de 1.000 colaboradores",
]

VALID_ANNUAL_REVENUES = [
    "Ate R$ 10 milhoes",
    "R$ 10-30 milhoes",
    "R$ 30-100 milhoes",
    "R$ 100-300 milhoes",
    "R$ 300 milhoes - R$ 1 bilhao",
    "Acima de R$ 1 bilhao",
]

# Sugestões de detalhe do canal (não bloqueante — usuário pode digitar livremente)
CHANNEL_DETAIL_SUGGESTIONS = [
    "Base - Resgate",
    "Base - Levantada de Mão",
    "Inbound - Site",
    "Inbound - LinkedIn Ads",
    "Inbound - Google Ads",
    "Inbound - Webinar",
    "Inbound - Formulário",
    "Outbound - Cold Call",
    "Outbound - LinkedIn",
    "Outbound - WhatsApp",
    "Outbound - Cold Email",
    "Indicação - Cliente",
    "Indicação - Parceiro",
    "Parcerias - Parceiro",
    "Eventos - Presencial",
    "Eventos - Online",
]

STATE_TO_UF = {
    "acre": "AC", "alagoas": "AL", "amapa": "AP", "amazonas": "AM",
    "bahia": "BA", "ceara": "CE", "distrito federal": "DF",
    "espirito santo": "ES", "goias": "GO", "maranhao": "MA",
    "mato grosso do sul": "MS", "mato grosso": "MT", "minas gerais": "MG",
    "para": "PA", "paraiba": "PB", "parana": "PR", "pernambuco": "PE",
    "piaui": "PI", "rio de janeiro": "RJ", "rio grande do norte": "RN",
    "rio grande do sul": "RS", "rondonia": "RO", "roraima": "RR",
    "santa catarina": "SC", "sao paulo": "SP", "sergipe": "SE",
    "tocantins": "TO",
}
VALID_UF = sorted(STATE_TO_UF.values())


# ─── Helpers ────────────────────────────────────────────────────────────────

def _strip_accents(text: str) -> str:
    return unicode_normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def _clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _normalize_uf(val) -> str | None:
    s = _clean(val)
    if not s:
        return None
    upper = s.upper().strip()
    if upper in set(VALID_UF):
        return upper
    return STATE_TO_UF.get(_strip_accents(s.lower()).strip())


def _normalize_channel(val) -> str | None:
    s = _clean(val)
    if not s:
        return None
    return VALID_CHANNELS.get(_strip_accents(s.lower()).strip())


def _normalize_deal_type(val) -> str | None:
    s = _clean(val)
    if not s:
        return None
    return VALID_DEAL_TYPES.get(_strip_accents(s.lower()).strip())


def _parse_date(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _clean_cnpj(val) -> str | None:
    if val is None:
        return None
    digits = "".join(c for c in str(val).strip() if c.isdigit())
    if not digits:
        return None
    return digits.zfill(14)


def _find_user(db: Session, name: str) -> int | None:
    if not name:
        return None
    name_norm = _strip_accents(name.lower())
    for user in db.query(User).filter(User.is_active == True).all():
        if not user.name:
            continue
        user_norm = _strip_accents(user.name.lower())
        significant = [w for w in user_norm.split() if len(w) > 3]
        if not significant:
            continue
        if all(w in name_norm for w in significant):
            return user.id
    return None


def _get_or_create_client(db: Session, row: dict) -> int | None:
    company_name = _clean(row.get("Nome da Empresa"))
    if not company_name:
        return None

    cnpj_digits = _clean_cnpj(row.get("CNPJ/CPF"))

    if cnpj_digits:
        for c in db.query(Client).filter(Client.is_deleted == False).all():
            if c.document:
                doc_digits = "".join(ch for ch in c.document if ch.isdigit()).zfill(14)
                if doc_digits == cnpj_digits:
                    return c.id

    if not cnpj_digits:
        existing = (
            db.query(Client)
            .filter(Client.name.ilike(company_name), Client.is_deleted == False)
            .first()
        )
        if existing:
            return existing.id

    new_client = Client(
        name=company_name,
        company_name=company_name,
        document=cnpj_digits,
        website=_clean(row.get("Website")),
        state=_normalize_uf(row.get("Estado (UF)")),
        city=_clean(row.get("Cidade")),
        cnae=_clean(row.get("CNAE")),
        sector=_clean(row.get("Setor / Segmento")),
        source="importacao",
        is_active=True,
        relationship_type="Lead",
        commercial_activity="Ativo",
    )
    db.add(new_client)
    db.flush()
    return new_client.id


def _get_or_create_person(db: Session, row: dict, client_id: int | None) -> int | None:
    name = _clean(row.get("Nome do Contato"))
    if not name:
        return None

    email = _clean(row.get("E-mail Comercial"))

    if email:
        existing = db.query(Person).filter(Person.email_commercial == email).first()
        if existing:
            return existing.id

    if client_id:
        existing = (
            db.query(Person)
            .filter(Person.name == name, Person.organization_id == client_id)
            .first()
        )
        if existing:
            return existing.id

    parts = name.strip().split()
    first_name = parts[0] if parts else name
    last_name = " ".join(parts[1:]) if len(parts) > 1 else ""

    new_person = Person(
        name=name,
        first_name=first_name,
        last_name=last_name,
        position=_clean(row.get("Cargo")),
        email_commercial=email,
        phone=_clean(row.get("Telefone Principal")),
        phone_whatsapp=_clean(row.get("WhatsApp")),
        phone_commercial=_clean(row.get("Telefone Comercial")),
        phone_alternative=_clean(row.get("Telefone Alternativo")),
        phone_extra1=_clean(row.get("Telefone Extra 1")),
        phone_extra2=_clean(row.get("Telefone Extra 2")),
        linkedin=_clean(row.get("LinkedIn")),
        organization_id=client_id,
        is_active=True,
    )
    db.add(new_person)
    db.flush()
    return new_person.id


def _create_card(
    db: Session,
    row: dict,
    client_id: int | None,
    person_id: int | None,
    sdr_id: int | None,
    vendor_id: int | None,
    position: int,
) -> int:
    title = _clean(row.get("Título do Card"))
    if not title:
        raise ValueError("Título do Card é obrigatório")

    # Compatibilidade com templates antigos
    value = 0.0
    valor_raw = row.get("Valor (R$)")
    if valor_raw is not None:
        try:
            value = float(str(valor_raw).replace(",", ".").strip())
        except (ValueError, TypeError):
            value = 0.0

    due_date = _parse_date(row.get("Data de Vencimento"))

    new_card = Card(
        title=title,
        description=_clean(row.get("Descrição")),
        list_id=TARGET_LIST_ID,
        client_id=client_id,
        person_id=person_id,
        sdr_id=sdr_id,
        assigned_to_id=vendor_id,
        value=value,
        due_date=due_date,
        is_won=0,
        acquisition_channel=_normalize_channel(row.get("Canal de Aquisição")),
        acquisition_channel_detail=_clean(row.get("Detalhe do Canal")),
        deal_type=_normalize_deal_type(row.get("Tipo de Negócio")) or "Nova Venda",
        origin=_clean(row.get("Origem")) or "Importação via planilha",
        position=position,
        prospection_entry_date=datetime.now(),
    )
    db.add(new_card)
    db.flush()

    history = CardListHistory(
        card_id=new_card.id,
        list_id=TARGET_LIST_ID,
        board_id=TARGET_BOARD_ID,
        entered_at=datetime.now(),
    )
    db.add(history)
    return new_card.id


# ─── Importação ─────────────────────────────────────────────────────────────

def process_import(
    db: Session,
    file_bytes: bytes,
    current_user_id: int,
    current_user_role: str = "",
) -> CardImportResponse:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val).strip() if val else f"__COL_{col}")

    col_index = {h: i + 1 for i, h in enumerate(headers)}

    def get_cell(row_num: int, col_name: str):
        idx = col_index.get(col_name)
        return ws.cell(row=row_num, column=idx).value if idx else None

    results: list[CardImportRowResult] = []
    created = 0
    errors = 0

    next_position = (
        db.query(Card)
        .filter(Card.list_id == TARGET_LIST_ID, Card.is_deleted == False)
        .count()
    )

    is_sdr = current_user_role == "sdr"
    is_salesperson = current_user_role == "salesperson"

    for row_num in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        row_data = {col_name: get_cell(row_num, col_name) for col_name in headers}
        title = _clean(row_data.get("Título do Card"))

        # Ignora linha de exemplo independentemente de onde estiver
        if title and title.startswith(EXEMPLO_PREFIX):
            continue

        if not title:
            results.append(CardImportRowResult(
                row=row_num, status="error", card_id=None, title=None,
                message="Título do Card é obrigatório",
            ))
            errors += 1
            continue

        try:
            # SDR: só pode definir seu próprio sdr_id, nunca vendor_id
            # Vendedor: só pode definir seu próprio vendor_id, nunca sdr_id
            if is_sdr:
                sdr_id = current_user_id
                vendor_id = None
            elif is_salesperson:
                vendor_id = current_user_id
                sdr_id = None
            else:
                sdr_name = _clean(row_data.get("SDR (nome exato)"))
                sdr_id = _find_user(db, sdr_name) if sdr_name else None
                vendor_name = _clean(row_data.get("Vendedor (nome exato)"))
                vendor_id = _find_user(db, vendor_name) if vendor_name else None

            client_id = _get_or_create_client(db, row_data)
            person_id = _get_or_create_person(db, row_data, client_id)
            card_id = _create_card(db, row_data, client_id, person_id, sdr_id, vendor_id, next_position)

            db.commit()
            next_position += 1
            created += 1

            results.append(CardImportRowResult(
                row=row_num, status="success", card_id=card_id, title=title,
                message="Card criado com sucesso",
            ))

        except Exception as e:
            db.rollback()
            errors += 1
            results.append(CardImportRowResult(
                row=row_num, status="error", card_id=None, title=title,
                message=str(e),
            ))

    return CardImportResponse(total=len(results), created=created, errors=errors, results=results)


# ─── Geração do template ─────────────────────────────────────────────────────

def generate_template(user_name: str = "", user_role: str = "") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importação"

    is_sdr = user_role == "sdr"
    is_salesperson = user_role == "salesperson"

    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CBD5E0"),
        right=Side(style="thin", color="CBD5E0"),
        bottom=Side(style="thin", color="CBD5E0"),
    )

    example_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    example_font = Font(color="856404", italic=True, size=10)
    example_title_font = Font(color="856404", italic=True, bold=True, size=10)

    locked_fill = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")
    locked_header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
    locked_font = Font(color="6B7280", italic=True, size=10)

    # ── Pré-preenchimento de exemplo com dados do usuário logado ─────────────
    example_values: dict[str, str] = {}
    for header, example, _, section, _ in TEMPLATE_COLUMNS:
        if header == "SDR (nome exato)":
            if is_sdr:
                example_values[header] = user_name or "Maria Santos"
            elif is_salesperson:
                example_values[header] = "N/A"
            else:
                example_values[header] = "Maria Santos"
        elif header == "Vendedor (nome exato)":
            if is_salesperson:
                example_values[header] = user_name or "João Silva"
            elif is_sdr:
                example_values[header] = "N/A"
            else:
                example_values[header] = "João Silva"
        else:
            example_values[header] = example

    example_values["Título do Card"] = f"{EXEMPLO_PREFIX} Proposta Empresa XYZ"

    # ── Mapa de índice de coluna pelo nome ───────────────────────────────────
    col_map = {h: i for i, (h, _, _, _, _) in enumerate(TEMPLATE_COLUMNS, 1)}

    def col_letter(col_name: str) -> str:
        return ws.cell(row=1, column=col_map[col_name]).column_letter

    def is_locked_col(header: str) -> bool:
        if is_sdr and header == "Vendedor (nome exato)":
            return True
        if is_salesperson and header == "SDR (nome exato)":
            return True
        return False

    # ── Linha 1: cabeçalhos ──────────────────────────────────────────────────
    for col_idx, (header, _, is_required, section, _) in enumerate(TEMPLATE_COLUMNS, 1):
        if is_locked_col(header):
            label = header + " (não aplicável)"
            fill = locked_header_fill
        else:
            label = header + (" *" if is_required else "")
            color = SECTION_COLORS[section]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # ── Linha 2: exemplo ─────────────────────────────────────────────────────
    for col_idx, (header, _, _, _, _) in enumerate(TEMPLATE_COLUMNS, 1):
        val = example_values[header]
        locked = is_locked_col(header)
        cell = ws.cell(row=2, column=col_idx, value=val)
        if locked:
            cell.fill = locked_fill
            cell.font = locked_font
        else:
            cell.fill = example_fill
            cell.font = example_title_font if col_idx == 1 else example_font
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border

    # ── Larguras das colunas ─────────────────────────────────────────────────
    col_widths = [
        30, 38, 20, 22, 32,              # Card (5)
        32, 20, 20, 25, 16, 10, 28, 28, 14,  # Empresa (9)
        28, 22, 28, 22, 20, 22, 22, 20, 20, 30,  # Contato (10)
        24, 24,                              # Responsáveis (2)
    ]
    for col_idx, width in enumerate(col_widths[:len(TEMPLATE_COLUMNS)], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    # ── Aba oculta com listas de validação ───────────────────────────────────
    lists_ws = wb.create_sheet("Listas")
    lists_ws.sheet_state = "hidden"

    canais = ["Inbound", "Outbound", "Indicacao", "Parcerias", "Eventos", "Base"]
    tipos = ["Nova Venda", "Cross Sell", "Up Sell"]

    for i, v in enumerate(canais, 1):
        lists_ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(tipos, 1):
        lists_ws.cell(row=i, column=2, value=v)
    for i, v in enumerate(VALID_UF, 1):
        lists_ws.cell(row=i, column=3, value=v)
    for i, v in enumerate(VALID_EMPLOYEE_COUNTS, 1):
        lists_ws.cell(row=i, column=4, value=v)
    for i, v in enumerate(VALID_ANNUAL_REVENUES, 1):
        lists_ws.cell(row=i, column=5, value=v)
    for i, v in enumerate(CHANNEL_DETAIL_SUGGESTIONS, 1):
        lists_ws.cell(row=i, column=6, value=v)

    def add_dv(formula: str, col_name: str, warning: bool = False):
        """warning=True permite digitar livremente além da lista (sugestões)."""
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="warning" if warning else "stop",
            error="Use um dos valores da lista." if not warning else "Valor não está na lista de sugestões. Você pode digitar livremente se necessário.",
            errorTitle="Valor inválido" if not warning else "Sugestão não encontrada",
        )
        ws.add_data_validation(dv)
        letter = col_letter(col_name)
        dv.add(f"{letter}3:{letter}10000")

    add_dv("Listas!$A$1:$A$6", "Canal de Aquisição")
    add_dv("Listas!$B$1:$B$3", "Tipo de Negócio")
    add_dv(f"Listas!$C$1:$C${len(VALID_UF)}", "Estado (UF)")
    add_dv(f"Listas!$D$1:$D${len(VALID_EMPLOYEE_COUNTS)}", "Nº de Funcionários")
    add_dv(f"Listas!$E$1:$E${len(VALID_ANNUAL_REVENUES)}", "Receita Anual")
    # Detalhe do Canal: lista de sugestões, mas permite texto livre (warning)
    add_dv(f"Listas!$F$1:$F${len(CHANNEL_DETAIL_SUGGESTIONS)}", "Detalhe do Canal", warning=True)

    # ── Colunas bloqueadas: DV que impede qualquer digitação ──────────────────
    for header, _, _, _, _ in TEMPLATE_COLUMNS:
        if is_locked_col(header):
            dv_lock = DataValidation(
                type="custom",
                formula1=f'"{col_letter(header)}2"',  # força célula vazia/N-A
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                error="Este campo não é aplicável ao seu perfil.",
                errorTitle="Coluna bloqueada",
            )
            ws.add_data_validation(dv_lock)
            letter = col_letter(header)
            dv_lock.add(f"{letter}3:{letter}10000")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
