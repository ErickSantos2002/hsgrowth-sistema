"""
Endpoints de Relatórios Customizados.
Cobre catálogo de campos, query engine e CRUD de relatórios salvos.

Todos os endpoints requerem autenticação e role admin ou manager.
"""
import io
import csv
from typing import Any, List
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session

from app.api.deps import get_db, get_current_active_user
from app.models.user import User
from app.services.custom_report_service import CustomReportService
from app.schemas.custom_report import (
    QueryRequest,
    QueryResponse,
    CustomReportCreate,
    CustomReportResponse,
    FieldCatalogResponse,
    DrillDownRequest,
    DrillDownResponse,
    ValidateFormulaRequest,
    ValidateFormulaResponse,
    CalculatedYFieldSchema,
)
from app.core.formula_evaluator import FormulaEvaluator
from app.schemas.report import PeriodEnum

router = APIRouter()


def _require_manager_or_admin(current_user: User) -> None:
    """
    Garante que o usuário tem role de admin ou manager.
    Lança 403 Forbidden caso contrário.
    """
    role_name = current_user.role.name if current_user.role else ""
    if role_name not in ("admin", "manager"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Apenas administradores e gerentes podem acessar os relatórios customizados",
        )


# ========================
# CATÁLOGO DE CAMPOS
# ========================

@router.get(
    "/fields",
    response_model=FieldCatalogResponse,
    summary="Catálogo de campos disponíveis",
    responses={
        200: {
            "description": "Catálogo retornado com sucesso",
        }
    },
)
async def get_report_fields(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    """
    Retorna o catálogo de campos disponíveis por fonte de dados.

    **Fontes:** cards, clients, persons, activities.
    Cada campo inclui: key, label, field_type, groupable, aggregatable.
    """
    _require_manager_or_admin(current_user)
    service = CustomReportService(db)
    return service.get_field_catalog()


# ========================
# QUERY ENGINE
# ========================

@router.post(
    "/query",
    response_model=QueryResponse,
    summary="Executa query de um gráfico",
    responses={
        200: {
            "description": "Dados do gráfico gerados com sucesso",
            "content": {
                "application/json": {
                    "example": {
                        "labels": ["Inbound", "Outbound", "Indicação"],
                        "values": [35.0, 28.0, 22.0],
                        "total": 85.0,
                    }
                }
            },
        }
    },
)
async def query_chart(
    request: QueryRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    """
    Executa a query de um gráfico customizado e retorna os dados para renderização.

    **Single-série** (1 campo Y): retorna `labels`, `values`, `total`.
    **Multi-série** (2–4 campos Y): retorna `labels`, `series[]`.

    O campo X define a dimensão de agrupamento (categorias ou datas).
    O campo Y define a métrica e a função de agregação (count, sum, avg, distinct_count).
    """
    _require_manager_or_admin(current_user)
    service = CustomReportService(db)
    return service.execute_query(request)


@router.post(
    "/calculated-fields/validate",
    response_model=ValidateFormulaResponse,
    summary="Valida uma fórmula de campo calculado",
    responses={
        200: {"description": "Resultado da validação retornado com sucesso"},
    },
)
async def validate_calculated_formula(
    request: ValidateFormulaRequest,
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Valida uma fórmula DAX-like para campo calculado.

    Verifica:
    - Sintaxe da fórmula (apenas operações aritméticas +, -, *, /)
    - Se todos os [field_key] referenciados existem em `available_keys`
    - Retorna a lista de dependências extraídas da fórmula

    **Exemplos válidos:**
    - `[won_count] / [count] * 100`
    - `[value] / [won_count]`
    - `([won_count] + [proposal_count]) / [count]`
    """
    _require_manager_or_admin(current_user)

    evaluator = FormulaEvaluator()

    # Valida a sintaxe da fórmula
    errors = evaluator.validate(request.formula)

    # Valida que todas as dependências são keys disponíveis para a fonte
    dependencies = evaluator.extract_dependencies(request.formula)
    invalid_keys = dependencies - set(request.available_keys)
    if invalid_keys:
        for key in sorted(invalid_keys):
            errors.append(
                f"Campo '[{key}]' não existe ou não é agregável na fonte '{request.source}'."
            )

    return ValidateFormulaResponse(
        is_valid=len(errors) == 0,
        errors=errors,
        dependencies=sorted(dependencies),
    )


@router.post(
    "/drill-down",
    response_model=DrillDownResponse,
    summary="Detalha os cards de uma barra/fatia do gráfico",
)
async def drill_down_chart(
    request: DrillDownRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    """
    Retorna os cards que compõem a barra ou fatia clicada no gráfico.

    Recebe o mesmo eixo X do gráfico + o label da barra clicada.
    Quando o gráfico usa split_by, informar também o split_label (nome da série clicada).
    Limitado a 200 registros para não sobrecarregar a interface.
    """
    _require_manager_or_admin(current_user)
    service = CustomReportService(db)
    return service.execute_drill_down(request)


# ========================
# CRUD DE RELATÓRIOS
# ========================

@router.get(
    "/custom/{report_id}/export",
    summary="Exporta um relatório customizado em Excel ou CSV",
    responses={
        200: {"description": "Arquivo gerado com sucesso — retornado como download"},
        400: {"description": "Formato inválido"},
        404: {"description": "Relatório não encontrado"},
    },
)
async def export_custom_report(
    report_id: int,
    format: str = Query("excel", description="Formato de saída: 'excel' ou 'csv'"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    """
    Gera um arquivo de download com os dados de todos os gráficos do relatório.

    - **Excel**: uma aba por gráfico, com cabeçalho e linhas de dados.
    - **CSV**: arquivo único com seções separadas por gráfico (compatível com Excel via BOM UTF-8).

    Gráficos sem configuração de eixo X ou Y são ignorados silenciosamente.
    """
    _require_manager_or_admin(current_user)

    if format not in ("excel", "csv"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato inválido. Use 'excel' ou 'csv'.",
        )

    service = CustomReportService(db)
    report = service.get_report(report_id)  # Lança 404 se não encontrado

    charts = report.config.get("charts", [])
    # Campos calculados definidos no relatório (shared entre todos os gráficos)
    report_calculated_fields = report.config.get("calculated_fields") or []

    # Executa a query de cada gráfico configurado e acumula resultados
    chart_results: list[dict] = []
    for chart in charts:
        if not chart.get("x_field") or not chart.get("y_fields"):
            continue  # Ignora gráficos sem configuração mínima
        try:
            # Separa campos Y normais dos calculados para enviar no request correto
            raw_y_fields = chart.get("y_fields", [])
            normal_y_fields = [
                yf for yf in raw_y_fields if not yf.get("is_calculated")
            ]
            calc_y_fields = [
                yf for yf in raw_y_fields if yf.get("is_calculated")
            ]

            query_request = QueryRequest.model_validate({
                "x_field": chart["x_field"],
                "x_group_by": chart.get("x_group_by"),
                "y_fields": normal_y_fields,
                "calculated_y_fields": calc_y_fields or None,
                "calculated_fields": report_calculated_fields or None,
                "period": chart.get("period", PeriodEnum.THIS_MONTH),
                "start_date": chart.get("date_start"),
                "end_date": chart.get("date_end"),
                "split_by": chart.get("split_by"),
            })
            result = service.execute_query(query_request)
            chart_results.append({
                "title": chart.get("title") or "Gráfico",
                "result": result,
            })
        except Exception:
            # Erros individuais de gráfico não bloqueam a exportação dos demais
            continue

    # Sanitiza o nome do arquivo para uso seguro no header HTTP
    safe_name = "".join(c for c in report.name if c.isalnum() or c in " _-").strip() or "relatorio"

    # ─── Geração do Excel ────────────────────────────────────────────────────────
    if format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        wb.remove(wb.active)  # Remove a aba padrão vazia

        for item in chart_results:
            # Nome da aba: Excel limita a 31 caracteres
            sheet_title = item["title"][:31]
            ws = wb.create_sheet(title=sheet_title)
            result = item["result"]

            # ── Cabeçalho com destaque visual ────────────────────────────────
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            if result.series:
                # Gráfico multi-série: Label | Serie1 | Serie2 | ...
                headers = ["Label"] + [s.name for s in result.series]
            else:
                # Gráfico single-série: Label | Valor
                headers = ["Label", "Valor"]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # ── Linhas de dados ──────────────────────────────────────────────
            for row_idx, label in enumerate(result.labels, start=2):
                ws.cell(row=row_idx, column=1, value=label)
                if result.series:
                    for col_idx, serie in enumerate(result.series, start=2):
                        value = serie.values[row_idx - 2] if (row_idx - 2) < len(serie.values) else 0
                        ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    values = result.values or []
                    value = values[row_idx - 2] if (row_idx - 2) < len(values) else 0
                    ws.cell(row=row_idx, column=2, value=value)

            # Ajusta largura das colunas automaticamente
            for column in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in column), default=10)
                ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 60)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
        )

    # ─── Geração do CSV ──────────────────────────────────────────────────────────
    buffer_str = io.StringIO()
    writer = csv.writer(buffer_str)

    for idx, item in enumerate(chart_results):
        if idx > 0:
            writer.writerow([])  # Linha em branco separando gráficos

        writer.writerow([f"### {item['title']} ###"])
        result = item["result"]

        if result.series:
            writer.writerow(["Label"] + [s.name for s in result.series])
            for row_idx, label in enumerate(result.labels):
                row = [label] + [
                    s.values[row_idx] if row_idx < len(s.values) else 0
                    for s in result.series
                ]
                writer.writerow(row)
        else:
            writer.writerow(["Label", "Valor"])
            for label, value in zip(result.labels, result.values or []):
                writer.writerow([label, value])

    # BOM UTF-8 garante que o Excel abra o CSV com acentos corretamente
    content = "\ufeff" + buffer_str.getvalue()

    return Response(
        content=content.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
    )


@router.get(
    "/custom",
    response_model=List[CustomReportResponse],
    summary="Lista relatórios customizados",
    responses={
        200: {"description": "Lista retornada com sucesso"},
    },
)
async def list_custom_reports(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    """
    Lista todos os relatórios customizados da empresa, ordenados pelo mais recente.
    """
    _require_manager_or_admin(current_user)
    service = CustomReportService(db)
    return service.list_reports()


@router.post(
    "/custom",
    response_model=CustomReportResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Cria um novo relatório customizado",
)
async def create_custom_report(
    data: CustomReportCreate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    """
    Cria um novo relatório customizado e persiste no banco.

    O campo `config` deve conter o objeto `CustomReportConfig` serializado do frontend.
    """
    _require_manager_or_admin(current_user)
    service = CustomReportService(db)
    return service.create_report(data, user_id=current_user.id)


@router.get(
    "/custom/{report_id}",
    response_model=CustomReportResponse,
    summary="Busca relatório por ID",
    responses={
        404: {"description": "Relatório não encontrado"},
    },
)
async def get_custom_report(
    report_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    """Busca um relatório customizado pelo ID."""
    _require_manager_or_admin(current_user)
    service = CustomReportService(db)
    return service.get_report(report_id)


@router.put(
    "/custom/{report_id}",
    response_model=CustomReportResponse,
    summary="Atualiza um relatório customizado",
    responses={
        404: {"description": "Relatório não encontrado"},
    },
)
async def update_custom_report(
    report_id: int,
    data: CustomReportCreate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    """Atualiza nome e configuração de um relatório customizado."""
    _require_manager_or_admin(current_user)
    service = CustomReportService(db)
    return service.update_report(report_id, data)


@router.delete(
    "/custom/{report_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Exclui um relatório customizado",
    responses={
        204: {"description": "Relatório excluído com sucesso"},
        404: {"description": "Relatório não encontrado"},
    },
)
async def delete_custom_report(
    report_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> None:
    """Remove definitivamente um relatório customizado."""
    _require_manager_or_admin(current_user)
    service = CustomReportService(db)
    service.delete_report(report_id)
