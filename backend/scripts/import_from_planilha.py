"""
Script para importar dados da Planilha_Importacao_CRM_preenchida.xlsx e criar cards no CRM.

Regras de negócio aplicadas:
- Mapeamento de colunas feito por nome (não por posição), portanto colunas
  extras como LinkedIn são automaticamente ignoradas.
- Faixa de Funcionários / Faturamento fora do padrão CRM → importados como null.
- Estado como nome completo → normalizado para sigla UF; se não reconhecido → null.
- Canal de Aquisição → normalizado para o valor exato esperado pelo CRM.
- Contatos 2 e 3 → salvos na descrição do card (o CRM vincula 1 pessoa por card).
- Atividades → criadas apenas se Ativ_Data estiver preenchida.
- Duplicatas por CNPJ → cliente reutilizado, sem criar duplicata.
"""

import sys
import os
import traceback
from datetime import datetime
from unicodedata import normalize as unicode_normalize

import openpyxl

# Adiciona o diretório raiz do backend ao PYTHONPATH
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from app.db.session import SessionLocal
from app.models.card import Card
from app.models.client import Client
from app.models.person import Person
from app.models.user import User
from app.models.activity import Activity
from app.models.card_list_history import CardListHistory

# ============================================================
# CONSTANTES
# ============================================================

# Lista alvo: "Prospecção" no Board 6
TARGET_LIST_ID = 23
TARGET_BOARD_ID = 6

# Valores válidos para employee_count (campo CRM)
VALID_EMPLOYEE_COUNTS = [
    "Ate 50 colaboradores",
    "51-100 colaboradores",
    "101-300 colaboradores",
    "301-600 colaboradores",
    "601-1.000 colaboradores",
    "Acima de 1.000 colaboradores",
]

# Valores válidos para annual_revenue (campo CRM)
VALID_ANNUAL_REVENUES = [
    "Ate R$ 10 milhoes",
    "R$ 10-30 milhoes",
    "R$ 30-100 milhoes",
    "R$ 100-300 milhoes",
    "R$ 300 milhoes - R$ 1 bilhao",
    "Acima de R$ 1 bilhao",
]

# Valores válidos para acquisition_channel
VALID_CHANNELS = {
    "inbound": "Inbound",
    "outbound": "Outbound",
    "indicacao": "Indicacao",
    "parcerias": "Parcerias",
    "eventos": "Eventos",
    "base": "Base",
}

# Valores válidos para deal_type
VALID_DEAL_TYPES = {
    "nova venda": "Nova Venda",
    "cross sell": "Cross Sell",
    "up sell": "Up Sell",
}

# Mapa de nome do estado por extenso para sigla UF
STATE_TO_UF = {
    "acre": "AC", "alagoas": "AL", "amapa": "AP", "amazonas": "AM",
    "bahia": "BA", "ceara": "CE", "distrito federal": "DF",
    "espirito santo": "ES", "goias": "GO", "maranhao": "MA",
    "mato grosso do sul": "MS", "mato grosso": "MT", "minas gerais": "MG",
    "para": "PA", "paraiba": "PB", "parana": "PR", "pernambuco": "PE",
    "piaui": "PI", "rio de janeiro": "RJ", "rio grande do norte": "RN",
    "rio grande do sul": "RS", "rondonia": "RO", "roraima": "RR",
    "santa catarina": "SC", "sao paulo": "SP", "sergipe": "SE",
    "tocantins": "TO",
}

VALID_UF = set(STATE_TO_UF.values())


# ============================================================
# HELPERS DE NORMALIZAÇÃO
# ============================================================

def strip_accents(text: str) -> str:
    """Remove acentos de uma string para facilitar comparações."""
    return unicode_normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def clean_str(val) -> str | None:
    """Converte valor de célula para string limpa, retorna None se vazio."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def normalize_uf(val) -> str | None:
    """Normaliza estado para sigla UF de 2 letras. Retorna None se não reconhecido."""
    s = clean_str(val)
    if not s:
        return None
    upper = s.upper().strip()
    # Já é uma UF válida de 2 letras
    if upper in VALID_UF:
        return upper
    # Tenta por nome completo (ex: "SAO PAULO" → "SP")
    normalized = strip_accents(s.lower()).strip()
    return STATE_TO_UF.get(normalized)


def normalize_channel(val) -> str | None:
    """Normaliza canal de aquisição para o valor exato aceito pelo CRM."""
    s = clean_str(val)
    if not s:
        return None
    key = strip_accents(s.lower()).strip()
    return VALID_CHANNELS.get(key)


def normalize_deal_type(val) -> str | None:
    """Normaliza tipo de negócio para o valor exato aceito pelo CRM."""
    s = clean_str(val)
    if not s:
        return None
    key = strip_accents(s.lower()).strip()
    return VALID_DEAL_TYPES.get(key)


def normalize_employee_count(val) -> str | None:
    """
    Valida faixa de funcionários contra os valores aceitos pelo CRM.
    Retorna None se o valor estiver fora do padrão (sem tentar adivinhar).
    """
    s = clean_str(val)
    if not s:
        return None
    s_norm = strip_accents(s.lower()).strip()
    for valid in VALID_EMPLOYEE_COUNTS:
        if strip_accents(valid.lower()) == s_norm:
            return valid
    return None  # Fora do padrão — ignora


def normalize_annual_revenue(val) -> str | None:
    """
    Valida faixa de faturamento contra os valores aceitos pelo CRM.
    Retorna None se o valor estiver fora do padrão.
    """
    s = clean_str(val)
    if not s:
        return None
    s_norm = strip_accents(s.lower()).strip()
    for valid in VALID_ANNUAL_REVENUES:
        if strip_accents(valid.lower()) == s_norm:
            return valid
    return None  # Fora do padrão — ignora


def parse_date(val) -> datetime | None:
    """Converte múltiplos formatos de data para datetime."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def clean_phone(val) -> str | None:
    """Converte número de telefone (inteiro ou string) para string limpa."""
    if val is None:
        return None
    return str(val).strip() or None


def extract_cnae_code(val) -> str | None:
    """
    Extrai somente o código CNAE (parte numérica) da string completa.
    Ex: "4930202 - TRANSPORTE RODOVIARIO..." → "4930202"
    A coluna cnae no banco é VARCHAR(20), então não cabe a descrição inteira.
    """
    s = clean_str(val)
    if not s:
        return None
    # Se tiver " - ", pega tudo antes disso (o código)
    if " - " in s:
        code = s.split(" - ")[0].strip()
    else:
        code = s.strip()
    # Garante que não passa de 20 chars (segurança extra)
    return code[:20]


def find_sdr_user(db, sdr_name: str) -> int | None:
    """
    Busca usuário no banco pelo nome informado na planilha.
    Exige que TODAS as palavras significativas (>3 chars) do nome do usuário
    estejam presentes no nome buscado — evita falsos positivos por sobrenome comum
    (ex: 'Sandra Silva' não bate em 'Cláudia Silva' porque 'sandra' não está lá).
    """
    if not sdr_name:
        return None
    name_norm = strip_accents(sdr_name.lower())
    users = db.query(User).filter(User.is_active == True).all()
    for user in users:
        if not user.name:
            continue
        user_norm = strip_accents(user.name.lower())
        # Palavras significativas do nome do usuário cadastrado (mais de 3 letras)
        significant_words = [word for word in user_norm.split() if len(word) > 3]
        if not significant_words:
            continue
        # Todas as palavras significativas do usuário devem estar no nome da planilha
        all_match = all(word in name_norm for word in significant_words)
        if all_match:
            return user.id
    return None


# ============================================================
# LEITOR DE COLUNAS POR NOME
# ============================================================

class RowReader:
    """
    Encapsula a leitura de uma linha da planilha por nome de coluna.
    O mapeamento é feito sem acentos e case-insensitive para robustez.
    """

    def __init__(self, ws, headers: list[str]):
        self._ws = ws
        # Cria índice: versão normalizada do header → número da coluna (1-based)
        self._index: dict[str, int] = {}
        for col_idx, h in enumerate(headers, 1):
            key = strip_accents(h.lower().strip())
            self._index[key] = col_idx

    def get(self, row: int, column_name: str):
        """Retorna o valor da célula pelo nome da coluna (tolerante a acentos e case)."""
        key = strip_accents(column_name.lower().strip())
        col_idx = self._index.get(key)
        if col_idx is None:
            return None
        return self._ws.cell(row=row, column=col_idx).value

    def has_column(self, column_name: str) -> bool:
        key = strip_accents(column_name.lower().strip())
        return key in self._index


# ============================================================
# FUNÇÕES DE CRIAÇÃO NO BANCO
# ============================================================

def get_or_create_client(db, reader: RowReader, row: int) -> int | None:
    """
    Busca cliente pelo CNPJ (somente dígitos) ou cria um novo.
    Razão Social é obrigatória — se ausente, retorna None.
    """
    razao_social = clean_str(reader.get(row, "Razão Social *"))
    if not razao_social:
        return None

    cnpj_raw = clean_str(reader.get(row, "CNPJ *"))
    cnpj_digits = "".join(c for c in cnpj_raw if c.isdigit()) if cnpj_raw else None

    # Tenta localizar cliente existente pelo CNPJ
    if cnpj_digits and len(cnpj_digits) == 14:
        existing_clients = db.query(Client).filter(Client.is_deleted == False).all()
        for c in existing_clients:
            if c.document:
                doc_digits = "".join(ch for ch in c.document if ch.isdigit())
                if doc_digits == cnpj_digits:
                    print(f"    Cliente ja existe (CNPJ ...{cnpj_digits[-4:]}): {c.name}")
                    return c.id

    # Cria novo cliente com os dados da planilha
    new_client = Client(
        name=razao_social,
        company_name=clean_str(reader.get(row, "Nome Fantasia")) or razao_social,
        document=cnpj_raw,
        website=clean_str(reader.get(row, "Site")),
        state=normalize_uf(reader.get(row, "Estado (UF)")),
        city=clean_str(reader.get(row, "Cidade")),
        address=clean_str(reader.get(row, "Endereço Completo")),
        cnae=extract_cnae_code(reader.get(row, "CNAE Principal")),
        employee_count=normalize_employee_count(reader.get(row, "Faixa de Funcionários")),
        annual_revenue=normalize_annual_revenue(reader.get(row, "Faixa de Faturamento")),
        phone=clean_phone(reader.get(row, "Fone1_Empresa")),
        email=clean_str(reader.get(row, "Email1_Empresa")),
        source="planilha_sdr",
        is_active=True,
    )
    db.add(new_client)
    db.flush()
    return new_client.id


def get_or_create_person(
    db, reader: RowReader, row: int, client_id: int | None
) -> int | None:
    """
    Busca ou cria a pessoa referente ao contato principal da linha.
    Se Nome_Contato1 não estiver preenchido, retorna None sem erro.
    """
    name = clean_str(reader.get(row, "Nome_Contato1 *"))
    if not name:
        return None

    # Evita duplicata: mesmo nome vinculado ao mesmo cliente
    if client_id:
        existing = (
            db.query(Person)
            .filter(Person.name == name, Person.organization_id == client_id)
            .first()
        )
        if existing:
            return existing.id

    # Divide nome em first/last para os campos separados do modelo
    parts = name.strip().split()
    first_name = parts[0] if parts else name
    last_name = " ".join(parts[1:]) if len(parts) > 1 else ""

    new_person = Person(
        name=name,
        first_name=first_name,
        last_name=last_name,
        position=clean_str(reader.get(row, "Cargo_Contato1")),
        phone=clean_phone(reader.get(row, "Tel_Principal_Contato1")),
        phone_whatsapp=clean_phone(reader.get(row, "Tel_WhatsApp_Contato1")),
        phone_commercial=clean_phone(reader.get(row, "Tel_Comercial_Contato1")),
        email=clean_str(reader.get(row, "Email_Contato1")),
        organization_id=client_id,
        is_active=True,
    )
    db.add(new_person)
    db.flush()
    return new_person.id


def build_description(reader: RowReader, row: int) -> str | None:
    """
    Monta a descrição do card com:
    1. Obs_Gerais_Card (campo principal de anotações do SDR)
    2. Dados dos contatos 2 e 3 (que não são vinculados como pessoa)
    3. Anotações livres
    """
    parts = []

    # Observações gerais do SDR
    obs = clean_str(reader.get(row, "Obs_Gerais_Card"))
    if obs:
        parts.append(obs)

    # Contato 2
    nome2 = clean_str(reader.get(row, "Nome_Contato2"))
    if nome2:
        lines = [f"Contato 2: {nome2}"]
        cargo2 = clean_str(reader.get(row, "Cargo_Contato2"))
        tel2 = clean_phone(reader.get(row, "Tel_Contato2"))
        email2 = clean_str(reader.get(row, "Email_Contato2"))
        if cargo2:
            lines.append(f"  Cargo: {cargo2}")
        if tel2:
            lines.append(f"  Tel: {tel2}")
        if email2:
            lines.append(f"  Email: {email2}")
        parts.append("\n".join(lines))

    # Contato 3
    nome3 = clean_str(reader.get(row, "Nome_Contato3"))
    if nome3:
        lines = [f"Contato 3: {nome3}"]
        cargo3 = clean_str(reader.get(row, "Cargo_Contato3"))
        tel3 = clean_phone(reader.get(row, "Tel_Contato3"))
        email3 = clean_str(reader.get(row, "Email_Contato3"))
        if cargo3:
            lines.append(f"  Cargo: {cargo3}")
        if tel3:
            lines.append(f"  Tel: {tel3}")
        if email3:
            lines.append(f"  Email: {email3}")
        parts.append("\n".join(lines))

    # Anotações livres
    anotacoes = clean_str(reader.get(row, "Anotacoes"))
    if anotacoes:
        parts.append(f"Anotações:\n{anotacoes}")

    return "\n\n".join(parts) if parts else None


def create_card(
    db,
    reader: RowReader,
    row: int,
    client_id: int | None,
    person_id: int | None,
    sdr_id: int | None,
    position: int,
) -> int | None:
    """Cria o card na lista Prospecção com todos os campos mapeados."""
    razao_social = clean_str(reader.get(row, "Razão Social *"))
    nome_fantasia = clean_str(reader.get(row, "Nome Fantasia"))
    # Título do card: prefere Nome Fantasia, usa Razão Social como fallback
    title = nome_fantasia or razao_social
    if not title:
        return None

    # Valor estimado: converte para float, descarta se inválido
    valor_raw = reader.get(row, "Valor_Estimado (R$)")
    value = 0.0
    if valor_raw is not None:
        try:
            value = float(str(valor_raw).replace(",", ".").strip())
        except (ValueError, TypeError):
            value = 0.0

    # Data de criação do card: prioriza "Data Criação*" (preenchida pelo SDR),
    # cai em "Data_Prospecao" como fallback, e por último usa hoje
    parsed_date = parse_date(reader.get(row, "Data Criação*")) or parse_date(reader.get(row, "Data_Prospecao"))
    if parsed_date and parsed_date > datetime.now():
        print(f"    AVISO   : Data '{parsed_date.date()}' está no futuro — usando data de hoje")
        parsed_date = None
    created_at = parsed_date or datetime.now()

    new_card = Card(
        title=title,
        list_id=TARGET_LIST_ID,
        client_id=client_id,
        person_id=person_id,
        sdr_id=sdr_id,
        assigned_to_id=None,  # Vendedor fica vazio — só o SDR é atribuído
        value=value,
        is_won=0,  # Card aberto
        acquisition_channel=normalize_channel(reader.get(row, "Canal_Aquisicao")),
        acquisition_channel_detail=clean_str(reader.get(row, "Canal_Aquisicao_Detalhe")),
        deal_type=normalize_deal_type(reader.get(row, "Tipo_Negocio")),
        description=build_description(reader, row),
        position=position,
        created_at=created_at,
        contact_info={
            "source": "planilha_sdr",
            "imported_at": datetime.now().isoformat(),
        },
    )
    db.add(new_card)
    db.flush()

    # Registra entrada no histórico de listas (tracking de board)
    # e preenche prospection_entry_date — replicando o que card_service faz
    new_card.prospection_entry_date = created_at
    history_entry = CardListHistory(
        card_id=new_card.id,
        list_id=TARGET_LIST_ID,
        board_id=TARGET_BOARD_ID,
        entered_at=created_at,
    )
    db.add(history_entry)

    return new_card.id


def create_activities(
    db, reader: RowReader, row: int, card_id: int, sdr_id: int | None
) -> int:
    """
    Cria até 3 atividades vinculadas ao card.
    A atividade só é criada se Ativ_Data estiver preenchida.
    Retorna o número de atividades criadas.
    """
    count = 0
    for i in range(1, 4):
        activity_date = parse_date(reader.get(row, f"Ativ{i}_Data"))
        if not activity_date:
            continue  # Sem data → pula esta atividade

        activity_type = clean_str(reader.get(row, f"Ativ{i}_Tipo")) or "Outro"
        description = clean_str(reader.get(row, f"Ativ{i}_Resultado")) or ""

        activity = Activity(
            card_id=card_id,
            user_id=sdr_id,
            activity_type=activity_type,
            description=description,
            activity_metadata={},
            created_at=activity_date,
        )
        db.add(activity)
        count += 1

    return count


# ============================================================
# IMPORTAÇÃO PRINCIPAL
# ============================================================

def import_from_sheet():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sheet_path = os.path.join(script_dir, "Planilha_Importacao_CRM_JOAO_VICTOR.xlsx")

    print("=" * 70)
    print("IMPORTACAO DA PLANILHA SDR PARA CARDS")
    print("=" * 70)
    print(f"Arquivo : {sheet_path}")
    print(f"Lista   : Prospecção (id={TARGET_LIST_ID}, board={TARGET_BOARD_ID})")

    if not os.path.exists(sheet_path):
        print(f"\nERRO: Arquivo não encontrado.")
        sys.exit(1)

    # Carrega planilha
    wb = openpyxl.load_workbook(sheet_path, data_only=True)
    ws = wb["Importação_CRM"]

    # Lê cabeçalhos da linha 3 (linha de nomes das colunas)
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        headers.append(str(val).strip() if val else f"__COL_{col}")

    reader = RowReader(ws, headers)

    # Conecta ao banco
    db = SessionLocal()
    stats = {
        "total": 0,
        "imported": 0,
        "errors": 0,
        "clients_created": 0,
        "persons_created": 0,
        "activities_created": 0,
        "warnings": [],
    }

    try:
        # Calcula próxima posição disponível na lista alvo
        next_position = (
            db.query(Card)
            .filter(Card.list_id == TARGET_LIST_ID, Card.is_deleted == False)
            .count()
        )

        print(f"\nIniciando importação a partir da posição {next_position}...\n")
        print("-" * 70)

        for row_num in range(5, ws.max_row + 1):
            # Pula linhas sem CNPJ (considera a linha vazia)
            cnpj_val = reader.get(row_num, "CNPJ *")
            if not cnpj_val:
                continue

            razao_social = clean_str(reader.get(row_num, "Razão Social *")) or "(sem nome)"
            stats["total"] += 1
            print(f"\n[{stats['total']}] Linha {row_num}: {razao_social}")

            try:
                # --- SDR ---
                sdr_name = clean_str(reader.get(row_num, "SDR_Responsavel *"))
                sdr_id = find_sdr_user(db, sdr_name)
                if sdr_id:
                    print(f"    SDR     : {sdr_name} -> id={sdr_id}")
                else:
                    msg = f"Linha {row_num}: SDR '{sdr_name}' nao encontrado"
                    print(f"    AVISO   : {msg}")
                    stats["warnings"].append(msg)

                # --- Cliente ---
                client_id_before = db.query(Client).count()
                client_id = get_or_create_client(db, reader, row_num)
                if client_id and db.query(Client).count() > client_id_before:
                    stats["clients_created"] += 1
                    print(f"    Cliente : criado (id={client_id})")
                elif client_id:
                    print(f"    Cliente : reutilizado (id={client_id})")
                else:
                    print(f"    AVISO   : Razão Social ausente — cliente nao criado")

                # --- Pessoa (contato principal) ---
                person_id_before = db.query(Person).count()
                person_id = get_or_create_person(db, reader, row_num, client_id)
                if person_id and db.query(Person).count() > person_id_before:
                    stats["persons_created"] += 1
                    print(f"    Pessoa  : criada (id={person_id})")
                elif person_id:
                    print(f"    Pessoa  : reutilizada (id={person_id})")
                else:
                    print(f"    Pessoa  : nao criada (Nome_Contato1 ausente)")

                # --- Card ---
                card_id = create_card(
                    db, reader, row_num, client_id, person_id, sdr_id, next_position
                )
                if not card_id:
                    print(f"    ERRO    : Nao foi possivel criar o card (titulo ausente)")
                    stats["errors"] += 1
                    db.rollback()
                    continue

                print(f"    Card    : criado (id={card_id})")

                # --- Atividades ---
                n_activities = create_activities(db, reader, row_num, card_id, sdr_id)
                if n_activities:
                    stats["activities_created"] += n_activities
                    print(f"    Ativids : {n_activities} criadas")

                db.commit()
                next_position += 1
                stats["imported"] += 1

            except Exception as e:
                db.rollback()
                stats["errors"] += 1
                print(f"    ERRO    : {e}")
                traceback.print_exc()

    finally:
        db.close()

    # Relatório final
    print("\n" + "=" * 70)
    print("RELATORIO FINAL")
    print("=" * 70)
    print(f"  Linhas processadas    : {stats['total']}")
    print(f"  Cards importados      : {stats['imported']}")
    print(f"  Clientes criados      : {stats['clients_created']}")
    print(f"  Pessoas criadas       : {stats['persons_created']}")
    print(f"  Atividades criadas    : {stats['activities_created']}")
    print(f"  Erros                 : {stats['errors']}")
    if stats["warnings"]:
        print(f"\n  Avisos ({len(stats['warnings'])}):")
        for w in stats["warnings"]:
            print(f"    - {w}")
    print("=" * 70)


if __name__ == "__main__":
    try:
        import_from_sheet()
    except KeyboardInterrupt:
        print("\n\nImportacao cancelada pelo usuario.")
        sys.exit(1)
    except Exception as e:
        print(f"\nERRO CRITICO: {e}")
        traceback.print_exc()
        sys.exit(1)
