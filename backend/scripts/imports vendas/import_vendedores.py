"""
Script para importar dados da Planilha_Importacao_Vendedores_LoteN.xlsx e criar cards no CRM.

Diferenças em relação ao import_from_planilha.py (SDR):
- O campo SDR_Responsavel contém o nome do VENDEDOR, mapeado para assigned_to_id.
- sdr_id fica None — esses leads são de responsabilidade dos vendedores, não dos SDRs.
- source = "planilha_vendedor"
- Aba da planilha: "Importacao"
- Dados começam na linha 4 (não linha 5)

Uso:
    cd backend
    python "scripts/imports vendas/import_vendedores.py" "scripts/imports vendas/Planilha_Importacao_Vendedores_Lote1.xlsx"
    python "scripts/imports vendas/import_vendedores.py" "scripts/imports vendas/Planilha_Importacao_Vendedores_Lote2.xlsx"

ATENÇÃO: rode cada comando UMA ÚNICA VEZ. Rodar duas vezes cria cards duplicados.
"""

import sys
import os
import traceback
from datetime import datetime
from unicodedata import normalize as unicode_normalize

import openpyxl

# Adiciona o diretório raiz do backend ao PYTHONPATH
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

from app.db.session import SessionLocal
from app.models.card import Card
from app.models.client import Client
from app.models.person import Person
from app.models.user import User
from app.models.card_list_history import CardListHistory

# ============================================================
# CONSTANTES
# ============================================================

# Lista alvo: "Lead Novo" no Board 6 (Prospecção)
TARGET_LIST_ID  = 22
TARGET_BOARD_ID = 6

VALID_EMPLOYEE_COUNTS = [
    "Ate 50 colaboradores",
    "51-100 colaboradores",
    "101-300 colaboradores",
    "301-600 colaboradores",
    "601-1.000 colaboradores",
    "Acima de 1.000 colaboradores",
]

VALID_ANNUAL_REVENUES = [
    "Ate R$ 10 milhoes",
    "R$ 10-30 milhoes",
    "R$ 30-100 milhoes",
    "R$ 100-300 milhoes",
    "R$ 300 milhoes - R$ 1 bilhao",
    "Acima de R$ 1 bilhao",
]

VALID_CHANNELS = {
    "inbound": "Inbound",
    "outbound": "Outbound",
    "indicacao": "Indicacao",
    "parcerias": "Parcerias",
    "eventos": "Eventos",
    "base": "Base",
}

STATE_TO_UF = {
    "acre": "AC", "alagoas": "AL", "amapa": "AP", "amazonas": "AM",
    "bahia": "BA", "ceara": "CE", "distrito federal": "DF",
    "espirito santo": "ES", "goias": "GO", "maranhao": "MA",
    "mato grosso do sul": "MS", "mato grosso": "MT", "minas gerais": "MG",
    "para": "PA", "paraiba": "PB", "parana": "PR", "pernambuco": "PE",
    "piaui": "PI", "rio de janeiro": "RJ", "rio grande do norte": "RN",
    "rio grande do sul": "RS", "rondonia": "RO", "roraima": "RR",
    "santa catarina": "SC", "sao paulo": "SP", "sergipe": "SE",
    "tocantins": "TO",
}

VALID_UF = set(STATE_TO_UF.values())

# Mapa de nome abreviado (planilha) → ID do vendedor no banco.
# Necessário quando o nome na planilha não bate automaticamente com o nome completo.
VENDOR_NAME_MAP: dict[str, int] = {
    "sandra":   5,   # Sandra Silva
    "adriana":  3,   # Adriana Oliveira
    "gislayne": 2,   # Gislayne Nunes
    "eduardo":  4,   # Eduardo Luna
}


# ============================================================
# HELPERS DE NORMALIZAÇÃO
# ============================================================

def strip_accents(text: str) -> str:
    return unicode_normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def clean_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def normalize_uf(val) -> str | None:
    s = clean_str(val)
    if not s:
        return None
    upper = s.upper().strip()
    if upper in VALID_UF:
        return upper
    normalized = strip_accents(s.lower()).strip()
    return STATE_TO_UF.get(normalized)


def normalize_channel(val) -> str | None:
    s = clean_str(val)
    if not s:
        return None
    key = strip_accents(s.lower()).strip()
    return VALID_CHANNELS.get(key)


def normalize_employee_count(val) -> str | None:
    s = clean_str(val)
    if not s:
        return None
    s_norm = strip_accents(s.lower()).strip()
    for valid in VALID_EMPLOYEE_COUNTS:
        if strip_accents(valid.lower()) == s_norm:
            return valid
    return None


def normalize_annual_revenue(val) -> str | None:
    s = clean_str(val)
    if not s:
        return None
    s_norm = strip_accents(s.lower()).strip()
    for valid in VALID_ANNUAL_REVENUES:
        if strip_accents(valid.lower()) == s_norm:
            return valid
    return None


def clean_phone(val) -> str | None:
    if val is None:
        return None
    return str(val).strip() or None


def clean_cnpj(val) -> str | None:
    if val is None:
        return None
    digits = "".join(c for c in str(val).strip() if c.isdigit())
    if not digits:
        return None
    return digits.zfill(14)


def extract_cnae_code(val) -> str | None:
    s = clean_str(val)
    if not s:
        return None
    if " - " in s:
        code = s.split(" - ")[0].strip()
    else:
        code = s.strip()
    return code[:20]


def find_vendor_user(db, vendor_name: str) -> int | None:
    """
    Busca vendedor pelo nome informado na planilha.
    1. Verifica VENDOR_NAME_MAP para nomes abreviados.
    2. Fallback: busca por palavras significativas no banco.
    """
    if not vendor_name:
        return None
    name_norm = strip_accents(vendor_name.lower().strip())

    if name_norm in VENDOR_NAME_MAP:
        return VENDOR_NAME_MAP[name_norm]

    users = db.query(User).filter(User.is_active == True).all()
    for user in users:
        if not user.name:
            continue
        user_norm = strip_accents(user.name.lower())
        significant_words = [w for w in user_norm.split() if len(w) > 3]
        if not significant_words:
            continue
        if all(w in name_norm for w in significant_words):
            return user.id
    return None


# ============================================================
# LEITOR DE COLUNAS POR NOME
# ============================================================

class RowReader:
    def __init__(self, ws, headers: list[str]):
        self._ws = ws
        self._index: dict[str, int] = {}
        for col_idx, h in enumerate(headers, 1):
            key = strip_accents(h.lower().strip())
            self._index[key] = col_idx

    def get(self, row: int, column_name: str):
        key = strip_accents(column_name.lower().strip())
        col_idx = self._index.get(key)
        if col_idx is None:
            return None
        return self._ws.cell(row=row, column=col_idx).value


# ============================================================
# FUNÇÕES DE CRIAÇÃO NO BANCO
# ============================================================

def get_or_create_client(db, reader: RowReader, row: int) -> int | None:
    razao_social = clean_str(reader.get(row, "Razão Social *"))
    if not razao_social:
        return None

    cnpj_digits = clean_cnpj(reader.get(row, "CNPJ *"))

    if cnpj_digits:
        existing_clients = db.query(Client).filter(Client.is_deleted == False).all()
        for c in existing_clients:
            if c.document:
                doc_digits = "".join(ch for ch in c.document if ch.isdigit()).zfill(14)
                if doc_digits == cnpj_digits:
                    print(f"    Cliente : já existe (CNPJ ...{cnpj_digits[-4:]}): {c.name}")
                    return c.id

    new_client = Client(
        name=razao_social,
        company_name=clean_str(reader.get(row, "Nome Fantasia")) or razao_social,
        document=cnpj_digits,
        state=normalize_uf(reader.get(row, "Estado (UF)")),
        city=clean_str(reader.get(row, "Cidade")),
        address=clean_str(reader.get(row, "Endereço Completo")),
        cnae=extract_cnae_code(reader.get(row, "CNAE Principal")),
        employee_count=normalize_employee_count(reader.get(row, "Faixa de Funcionários")),
        annual_revenue=normalize_annual_revenue(reader.get(row, "Faixa de Faturamento")),
        source="planilha_vendedor",
        is_active=True,
    )
    db.add(new_client)
    db.flush()
    return new_client.id


def get_or_create_person(db, reader: RowReader, row: int, client_id: int | None) -> int | None:
    name = clean_str(reader.get(row, "Nome_Contato1 *"))
    if not name:
        return None

    if client_id:
        existing = db.query(Person).filter(
            Person.name == name, Person.organization_id == client_id
        ).first()
        if existing:
            return existing.id

    parts = name.strip().split()
    first_name = parts[0] if parts else name
    last_name = " ".join(parts[1:]) if len(parts) > 1 else ""

    new_person = Person(
        name=name,
        first_name=first_name,
        last_name=last_name,
        position=clean_str(reader.get(row, "Cargo_Contato1")),
        phone=clean_phone(reader.get(row, "Tel_Principal_Contato1")),
        phone_whatsapp=clean_phone(reader.get(row, "Tel_WhatsApp_Contato1")),
        phone_commercial=clean_phone(reader.get(row, "Tel_Comercial_Contato1")),
        email=clean_str(reader.get(row, "Email_Contato1")),
        organization_id=client_id,
        is_active=True,
    )
    db.add(new_person)
    db.flush()
    return new_person.id


def build_description(reader: RowReader, row: int) -> str | None:
    parts = []

    obs = clean_str(reader.get(row, "Obs_Gerais_Card"))
    if obs:
        parts.append(obs)

    for i, label in [(2, "Contato 2"), (3, "Contato 3")]:
        nome = clean_str(reader.get(row, f"Nome_Contato{i}"))
        if not nome:
            continue
        lines = [f"{label}: {nome}"]
        cargo = clean_str(reader.get(row, f"Cargo_Contato{i}"))
        tel   = clean_phone(reader.get(row, f"Tel_Contato{i}"))
        email = clean_str(reader.get(row, f"Email_Contato{i}"))
        if cargo: lines.append(f"  Cargo: {cargo}")
        if tel:   lines.append(f"  Tel: {tel}")
        if email: lines.append(f"  Email: {email}")
        parts.append("\n".join(lines))

    anotacoes = clean_str(reader.get(row, "Anotacoes"))
    if anotacoes:
        parts.append(f"Anotações:\n{anotacoes}")

    return "\n\n".join(parts) if parts else None


def create_card(
    db, reader: RowReader, row: int,
    client_id: int | None, person_id: int | None,
    vendor_id: int | None, position: int,
) -> int | None:
    razao_social = clean_str(reader.get(row, "Razão Social *"))
    nome_fantasia = clean_str(reader.get(row, "Nome Fantasia"))
    title = nome_fantasia or razao_social
    if not title:
        return None

    valor_raw = reader.get(row, "Valor_Estimado (R$)")
    value = 0.0
    if valor_raw is not None:
        try:
            value = float(str(valor_raw).replace(",", ".").strip())
        except (ValueError, TypeError):
            value = 0.0

    created_at = datetime.now()

    new_card = Card(
        title=title,
        list_id=TARGET_LIST_ID,
        client_id=client_id,
        person_id=person_id,
        sdr_id=None,               # Sem SDR — responsabilidade é do vendedor
        assigned_to_id=vendor_id,  # Vendedor responsável
        value=value,
        is_won=0,
        acquisition_channel=normalize_channel(reader.get(row, "Canal_Aquisicao")),
        acquisition_channel_detail=clean_str(reader.get(row, "Canal_Aquisicao_Detalhe")),
        origin="Importação via planilha",
        description=build_description(reader, row),
        position=position,
        created_at=created_at,
        contact_info={
            "source": "planilha_vendedor",
            "imported_at": datetime.now().isoformat(),
        },
    )
    db.add(new_card)
    db.flush()

    new_card.prospection_entry_date = created_at
    history_entry = CardListHistory(
        card_id=new_card.id,
        list_id=TARGET_LIST_ID,
        board_id=TARGET_BOARD_ID,
        entered_at=created_at,
    )
    db.add(history_entry)

    return new_card.id


# ============================================================
# IMPORTAÇÃO PRINCIPAL
# ============================================================

def import_from_sheet():
    if len(sys.argv) < 2:
        print("Uso: python import_vendedores.py <caminho_planilha.xlsx>")
        sys.exit(1)

    sheet_path = sys.argv[1]
    if not os.path.isabs(sheet_path):
        sheet_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", sheet_path)

    print("=" * 70)
    print("IMPORTACAO DE VENDEDORES PARA CARDS")
    print("=" * 70)
    print(f"Arquivo : {sheet_path}")
    print(f"Lista   : Lead Novo (id={TARGET_LIST_ID}, board={TARGET_BOARD_ID})")

    if not os.path.exists(sheet_path):
        print(f"\nERRO: Arquivo não encontrado: {sheet_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(sheet_path, data_only=True)
    ws = wb["Importacao"]

    # Cabeçalhos na linha 3
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        headers.append(str(val).strip() if val else f"__COL_{col}")

    reader = RowReader(ws, headers)
    db = SessionLocal()

    stats = {
        "total": 0, "imported": 0, "errors": 0,
        "clients_created": 0, "persons_created": 0, "warnings": [],
    }

    try:
        next_position = (
            db.query(Card)
            .filter(Card.list_id == TARGET_LIST_ID, Card.is_deleted == False)
            .count()
        )

        print(f"\nIniciando importação a partir da posição {next_position}...\n")
        print("-" * 70)

        # Dados começam na linha 4
        for row_num in range(4, ws.max_row + 1):
            cnpj_val = reader.get(row_num, "CNPJ *")
            if not cnpj_val:
                continue

            razao_social = clean_str(reader.get(row_num, "Razão Social *")) or "(sem nome)"
            stats["total"] += 1
            print(f"\n[{stats['total']}] Linha {row_num}: {razao_social}")

            try:
                # --- Vendedor (campo SDR_Responsavel reutilizado) ---
                vendor_name = clean_str(reader.get(row_num, "SDR_Responsavel *"))
                vendor_id = find_vendor_user(db, vendor_name) if vendor_name else None
                if vendor_name and vendor_id:
                    print(f"    Vendedor: {vendor_name} -> id={vendor_id}")
                elif vendor_name:
                    msg = f"Linha {row_num}: Vendedor '{vendor_name}' não encontrado"
                    print(f"    AVISO   : {msg}")
                    stats["warnings"].append(msg)

                # --- Cliente ---
                client_count_before = db.query(Client).count()
                client_id = get_or_create_client(db, reader, row_num)
                if client_id and db.query(Client).count() > client_count_before:
                    stats["clients_created"] += 1
                    print(f"    Cliente : criado (id={client_id})")
                elif client_id:
                    print(f"    Cliente : reutilizado (id={client_id})")
                else:
                    print(f"    AVISO   : Razão Social ausente — cliente não criado")

                # --- Pessoa (contato principal) ---
                person_count_before = db.query(Person).count()
                person_id = get_or_create_person(db, reader, row_num, client_id)
                if person_id and db.query(Person).count() > person_count_before:
                    stats["persons_created"] += 1
                    print(f"    Pessoa  : criada (id={person_id})")
                elif person_id:
                    print(f"    Pessoa  : reutilizada (id={person_id})")
                else:
                    print(f"    Pessoa  : não criada (Nome_Contato1 ausente)")

                # --- Card ---
                card_id = create_card(
                    db, reader, row_num, client_id, person_id, vendor_id, next_position
                )
                if not card_id:
                    print(f"    ERRO    : Não foi possível criar o card (título ausente)")
                    stats["errors"] += 1
                    db.rollback()
                    continue

                print(f"    Card    : criado (id={card_id})")

                db.commit()
                next_position += 1
                stats["imported"] += 1

            except Exception as e:
                db.rollback()
                stats["errors"] += 1
                print(f"    ERRO    : {e}")
                traceback.print_exc()

    finally:
        db.close()

    print("\n" + "=" * 70)
    print("RELATORIO FINAL")
    print("=" * 70)
    print(f"  Linhas processadas    : {stats['total']}")
    print(f"  Cards importados      : {stats['imported']}")
    print(f"  Clientes criados      : {stats['clients_created']}")
    print(f"  Pessoas criadas       : {stats['persons_created']}")
    print(f"  Erros                 : {stats['errors']}")
    if stats["warnings"]:
        print(f"\n  Avisos ({len(stats['warnings'])}):")
        for w in stats["warnings"]:
            print(f"    - {w}")
    print("=" * 70)


if __name__ == "__main__":
    try:
        import_from_sheet()
    except KeyboardInterrupt:
        print("\n\nImportação cancelada pelo usuário.")
        sys.exit(1)
    except Exception as e:
        print(f"\nERRO CRÍTICO: {e}")
        traceback.print_exc()
        sys.exit(1)
