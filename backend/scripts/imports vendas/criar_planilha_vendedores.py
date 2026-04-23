"""
Script: criar_planilha_vendedores.py
Converte Prospects_Distribuicao_Vendedores.xlsx para o formato de importação do CRM.
Cada aba (sandra, adriana, gislaine, eduardo) vira um bloco de linhas na planilha de saída,
com o nome do vendedor no campo SDR_Responsavel (usado para assigned_to_id na importação).

Uso:
    cd backend
    python "scripts/imports vendas/criar_planilha_vendedores.py"

Saída:
    scripts/imports vendas/Planilha_Importacao_Vendedores.xlsx
"""

import sys
import os
import openpyxl
from openpyxl import load_workbook
from copy import copy

sys.stdout.reconfigure(encoding="utf-8")

# ── Caminhos ──────────────────────────────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_FILE = os.path.join(BASE_DIR, "..", "imports", "Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx")

LOTES = [
    {
        "prospects": os.path.join(BASE_DIR, "Prospects_Distribuicao_Vendedores.xlsx"),
        "output":    os.path.join(BASE_DIR, "Planilha_Importacao_Vendedores_Lote1.xlsx"),
        "label":     "Lote 1",
    },
    {
        "prospects": os.path.join(BASE_DIR, "Prospects_Distribuicao_Vendedores_02.xlsx"),
        "output":    os.path.join(BASE_DIR, "Planilha_Importacao_Vendedores_Lote2.xlsx"),
        "label":     "Lote 2",
    },
]

# ── Mapeamento de abas → nome do vendedor no sistema ─────────────────────────
# Deve bater exatamente com o campo "name" do usuário no banco (ou um apelido
# configurado no SDR_NAME_MAP do import_from_planilha.py).
VENDEDOR_MAP = {
    "sandra":   "Sandra",
    "adriana":  "Adriana",
    "gislaine": "Gislayne",
    "eduardo":  "Eduardo",
}

# ── Faixas de funcionários (Qtd. numérica → label CRM) ───────────────────────
def faixa_funcionarios(qtd):
    if qtd is None:
        return None
    try:
        n = int(qtd)
    except (ValueError, TypeError):
        return None
    if n <= 50:
        return "Ate 50 colaboradores"
    if n <= 100:
        return "51-100 colaboradores"
    if n <= 300:
        return "101-300 colaboradores"
    if n <= 600:
        return "301-600 colaboradores"
    if n <= 1000:
        return "601-1.000 colaboradores"
    return "Acima de 1.000 colaboradores"

# ── Faixas de faturamento (valor numérico → label CRM) ───────────────────────
def faixa_faturamento(valor):
    if valor is None:
        return None
    try:
        v = float(valor)
    except (ValueError, TypeError):
        return None
    if v <= 10_000_000:
        return "Ate R$ 10 milhoes"
    if v <= 30_000_000:
        return "R$ 10-30 milhoes"
    if v <= 100_000_000:
        return "R$ 30-100 milhoes"
    if v <= 300_000_000:
        return "R$ 100-300 milhoes"
    if v <= 1_000_000_000:
        return "R$ 300 milhoes - R$ 1 bilhao"
    return "Acima de R$ 1 bilhao"

# ── Monta telefone juntando DDD + número ─────────────────────────────────────
def montar_fone(ddd, numero):
    if not numero:
        return None
    ddd = str(int(ddd)) if ddd else ""
    numero = str(numero).replace("-", "").replace(" ", "")
    return f"{ddd}{numero}" if ddd else numero

# ── Monta endereço completo ───────────────────────────────────────────────────
def montar_endereco(row):
    # Colunas: 15=Logradouro, 16=Nº, 17=Complemento, 18=Bairro, 19=Cidade, 20=UF, 21=CEP
    partes = []
    logradouro  = row[14]   # col 15
    numero      = row[15]   # col 16
    complemento = row[16]   # col 17
    bairro      = row[17]   # col 18
    cidade      = row[18]   # col 19
    uf          = row[19]   # col 20
    cep         = row[20]   # col 21

    if logradouro:
        partes.append(str(logradouro))
    if numero:
        partes.append(str(numero))
    if complemento:
        partes.append(str(complemento))
    if bairro:
        partes.append(str(bairro))
    if cidade and uf:
        partes.append(f"{cidade} - {uf}")
    elif cidade:
        partes.append(str(cidade))
    if cep:
        partes.append(str(cep))
    return ", ".join(partes) if partes else None

# ── Monta CNAE "codigo - descricao" ──────────────────────────────────────────
def montar_cnae(row):
    cod  = row[10]   # col 11
    desc = row[11]   # col 12
    if cod and desc:
        return f"{cod} - {desc}"
    if cod:
        return str(cod)
    return None

# ── Lê uma linha do Prospects e devolve lista de 51 valores (formato CRM) ────
def converter_linha(row, vendedor_nome):
    # row é 0-indexed (row[0] = col 1)
    out = [None] * 51

    # Empresa
    out[0]  = row[0]                                    # CNPJ
    out[1]  = row[1]                                    # Razão Social
    out[2]  = row[2]                                    # Nome Fantasia
    out[3]  = None                                      # Site (não tem na planilha)
    out[4]  = row[19]                                   # Estado (UF)
    out[5]  = row[18]                                   # Cidade
    out[6]  = montar_endereco(row)                      # Endereço Completo
    out[7]  = montar_cnae(row)                          # CNAE Principal
    out[8]  = faixa_funcionarios(row[6])                # Faixa de Funcionários
    out[9]  = faixa_faturamento(row[4])                 # Faixa de Faturamento

    # Telefones empresa — vazios (telefones vão para a pessoa de contato)
    out[10] = None                                      # Fone1_Empresa
    out[11] = None                                      # Fone2_Empresa
    out[12] = None                                      # Fone3_Empresa

    # E-mails empresa — vazios (e-mails vão para a pessoa de contato)
    out[13] = None                                      # Email1_Empresa
    out[14] = None                                      # Email2_Empresa
    out[15] = None                                      # Email3_Empresa

    # Contato principal (Sócio 1) com todos os telefones e o primeiro e-mail
    # Se não houver sócio cadastrado, usa a razão social como nome genérico
    out[16] = row[36] or f"Contato {row[1]}"           # Nome_Contato1 (Sócio 1)
    out[17] = row[37]                                   # Cargo_Contato1
    out[18] = None                                      # LinkedIn
    out[19] = montar_fone(row[21], row[22])             # Tel_Principal_Contato1 (Tel 1 da empresa)
    out[20] = montar_fone(row[24], row[25])             # Tel_WhatsApp_Contato1  (Tel 2 da empresa)
    out[21] = montar_fone(row[27], row[28])             # Tel_Comercial_Contato1 (Tel 3 da empresa)
    out[22] = row[33]                                   # Email_Contato1 (E-mail 1)

    # Contato 2 (Sócio 2) — vai para descrição do card, com seu e-mail
    out[23] = row[40]                                   # Nome_Contato2
    out[24] = row[41]                                   # Cargo_Contato2
    out[25] = None                                      # LinkedIn
    out[26] = montar_fone(row[42], row[43])             # Tel_Contato2
    out[27] = row[34]                                   # Email_Contato2 (E-mail 2)

    # Contato 3 (Sócio 3) — vai para descrição do card, com seu e-mail
    out[28] = row[44]                                   # Nome_Contato3
    out[29] = row[45]                                   # Cargo_Contato3
    out[30] = None                                      # LinkedIn
    out[31] = montar_fone(row[46], row[47])             # Tel_Contato3
    out[32] = row[35]                                   # Email_Contato3 (E-mail 3)

    # CRM
    out[33] = vendedor_nome                             # SDR_Responsavel (aqui = vendedor)
    out[34] = "Outbound"                                # Canal_Aquisicao
    out[35] = "Outbound - Lista fria"                   # Canal_Aquisicao_Detalhe
    out[36] = None                                      # Tipo_Negocio
    out[37] = None                                      # Valor_Estimado
    out[38] = None                                      # Data_Prospecao
    out[39] = None                                      # Obs_Gerais_Card

    # Atividades e anotações — vazios
    out[40] = out[41] = out[42] = None
    out[43] = out[44] = out[45] = None
    out[46] = out[47] = out[48] = None
    out[49] = None

    # Status importação
    out[50] = None                                      # será preenchido como "Importado" após import

    return out


# ── Processa um lote ──────────────────────────────────────────────────────────
def processar_lote(template_ws, lote: dict):
    print(f"\n{'='*60}")
    print(f"  {lote['label']}")
    print(f"  Origem : {lote['prospects']}")
    print(f"  Saída  : {lote['output']}")
    print(f"{'='*60}")

    prospects_wb = load_workbook(lote["prospects"])

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Importacao"

    # Copia as 3 linhas de cabeçalho do template
    for row_idx in range(1, 4):
        for col_idx in range(1, 52):
            src = template_ws.cell(row=row_idx, column=col_idx)
            dst = out_ws.cell(row=row_idx, column=col_idx)
            dst.value = src.value

    abas = [name for name in prospects_wb.sheetnames if name.lower() in VENDEDOR_MAP]
    data_row = 4
    total = 0

    for aba in abas:
        vendedor_nome = VENDEDOR_MAP[aba.lower()]
        ws = prospects_wb[aba]
        print(f"\nAba '{aba}' → vendedor: {vendedor_nome}")

        convertidos = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row[0] and not row[1]:
                continue
            valores = converter_linha(row, vendedor_nome)
            for col_idx, val in enumerate(valores, start=1):
                out_ws.cell(row=data_row, column=col_idx).value = val
            data_row += 1
            convertidos += 1

        print(f"  {convertidos} empresas convertidas")
        total += convertidos

    out_wb.save(lote["output"])
    print(f"\n✅ Gerado: {lote['output']}")
    print(f"   Total: {total} empresas | linhas 4 até {data_row - 1}")
    return total


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Lendo template: {TEMPLATE_FILE}")
    template_wb = load_workbook(TEMPLATE_FILE)
    template_ws = template_wb.active

    grand_total = 0
    for lote in LOTES:
        if not os.path.exists(lote["prospects"]):
            print(f"\n⚠️  Arquivo não encontrado, pulando: {lote['prospects']}")
            continue
        grand_total += processar_lote(template_ws, lote)

    print(f"\n{'='*60}")
    print(f"  Total geral: {grand_total} empresas em {len(LOTES)} lotes")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
