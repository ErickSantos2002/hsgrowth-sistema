"""
Script para gerar a Planilha_Importacao_CRM.xlsx
Planilha padronizada para SDRs preencherem dados de empresas, contatos, cards e atividades.

Regras de negócio aplicadas:
- Contato 2 e 3: mantidos para coleta de dados, mas o script de importação os salvará
  nas Anotações do card (o CRM vincula apenas 1 pessoa por card).
- Campos com lista fixa (Canal, Tipo Negócio, Tipo Atividade, Estado, Faixa Funcionários)
  bloqueiam entrada inválida com error_style="stop".
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


wb = openpyxl.Workbook()

# ============================================================
# ABA DE REFERÊNCIA - listas para dropdowns (oculta)
# ============================================================
ws_ref = wb.active
ws_ref.title = "Referência"
ws_ref.sheet_state = "hidden"

ref_data = {
    # Faixa de Funcionários - valores exatos do CRM (blueprintOptions.ts EMPLOYEE_COUNTS)
    "A": ("Faixa_Funcionarios", [
        "Ate 50 colaboradores",
        "51-100 colaboradores",
        "101-300 colaboradores",
        "301-600 colaboradores",
        "601-1.000 colaboradores",
        "Acima de 1.000 colaboradores",
    ]),
    # Canal de Aquisição - valores exatos do CRM (blueprintOptions.ts ACQUISITION_CHANNELS)
    "B": ("Canal_Aquisicao", [
        "Inbound", "Outbound", "Indicacao", "Parcerias", "Eventos", "Base",
    ]),
    # Canal de Aquisição - Detalhe (blueprintOptions.ts ACQUISITION_CHANNEL_DETAILS)
    "C": ("Canal_Aquisicao_Detalhe", [
        "Inbound - Conteudo",
        "Inbound - Trafego pago",
        "Inbound - SEO",
        "Inbound - Email marketing",
        "Inbound - Levantada de mao (site / WhatsApp / formulario)",
        "Outbound - Lista fria",
        "Outbound - LinkedIn",
        "Outbound - Cold email",
        "Outbound - Cold call",
        "Indicacao - Cliente",
        "Indicacao - Parceiro",
        "Indicacao - Network pessoal",
        "Parcerias - Co-marketing",
        "Parcerias - Integracao tecnologica",
        "Parcerias - Revenda",
        "Eventos - Feira",
        "Eventos - Webinar",
        "Eventos - Workshop",
        "Eventos - Meetup",
        "Base - Reativacao",
        "Base - Cross-sell",
        "Base - Up-sell",
    ]),
    # Tipo de Negócio - valores exatos do CRM (blueprintOptions.ts DEAL_TYPES)
    "D": ("Tipo_Negocio", [
        "Nova Venda", "Cross Sell", "Up Sell",
    ]),
    # Estado UF
    "E": ("Estado_UF", [
        "AC", "AL", "AM", "AP", "BA", "CE", "DF", "ES", "GO", "MA",
        "MG", "MS", "MT", "PA", "PB", "PE", "PI", "PR", "RJ", "RN",
        "RO", "RR", "RS", "SC", "SE", "SP", "TO",
    ]),
    # Tipo de Atividade
    "F": ("Tipo_Atividade", [
        "Ligacao", "WhatsApp", "Email", "Reuniao",
        "Visita presencial", "LinkedIn", "Fale Conosco / Site", "Outro",
    ]),
    # Faixa de Faturamento - valores exatos do CRM (blueprintOptions.ts ANNUAL_REVENUES)
    "G": ("Faixa_Faturamento", [
        "Ate R$ 10 milhoes",
        "R$ 10-30 milhoes",
        "R$ 30-100 milhoes",
        "R$ 100-300 milhoes",
        "R$ 300 milhoes - R$ 1 bilhao",
        "Acima de R$ 1 bilhao",
    ]),
}

for col_letter, (header, values) in ref_data.items():
    ws_ref[f"{col_letter}1"] = header
    for i, val in enumerate(values, 2):
        ws_ref[f"{col_letter}{i}"] = val


# ============================================================
# ABA PRINCIPAL
# ============================================================
ws = wb.create_sheet("Importação_CRM", 0)


# --- Helpers de estilo ---
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left_wrap():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


thin = Side(style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)


# --- Cores por seção ---
SECAO_COR_BG = {
    "DADOS DA EMPRESA":               "BDD7EE",
    "TELEFONES E EMAILS DA EMPRESA":  "DAEEF3",
    "CONTATO PRINCIPAL":              "FCE4D6",
    "CONTATO 2 (→ Anotações)":        "FDE9D9",
    "CONTATO 3 (→ Anotações)":        "FEF0E6",
    "CARD / NEGÓCIO NO CRM":          "E2DAEF",
    "ATIVIDADES REALIZADAS":          "FFF2CC",
    "ANOTAÇÕES":                      "EDEDED",
}

SECAO_COR_HEADER = {
    "DADOS DA EMPRESA":               "1F6091",
    "TELEFONES E EMAILS DA EMPRESA":  "17669A",
    "CONTATO PRINCIPAL":              "C55A11",
    "CONTATO 2 (→ Anotações)":        "C55A11",
    "CONTATO 3 (→ Anotações)":        "C55A11",
    "CARD / NEGÓCIO NO CRM":          "6A2D8F",
    "ATIVIDADES REALIZADAS":          "7F6000",
    "ANOTAÇÕES":                      "404040",
}


# --- Definição das colunas ---
# (seção, nome_coluna, largura, obrigatório)
columns = [
    # EMPRESA
    ("DADOS DA EMPRESA",              "CNPJ",                      18, True),
    ("DADOS DA EMPRESA",              "Razão Social",               35, True),
    ("DADOS DA EMPRESA",              "Nome Fantasia",              25, False),
    ("DADOS DA EMPRESA",              "Site",                       22, False),
    ("DADOS DA EMPRESA",              "Estado (UF)",                12, False),
    ("DADOS DA EMPRESA",              "Cidade",                     18, False),
    ("DADOS DA EMPRESA",              "Endereço Completo",          35, False),
    ("DADOS DA EMPRESA",              "CNAE Principal",             35, False),
    ("DADOS DA EMPRESA",              "Faixa de Funcionários",      22, False),
    ("DADOS DA EMPRESA",              "Faixa de Faturamento",       25, False),
    # FONES/EMAILS EMPRESA
    ("TELEFONES E EMAILS DA EMPRESA", "Fone1_Empresa",              22, False),
    ("TELEFONES E EMAILS DA EMPRESA", "Fone2_Empresa",              22, False),
    ("TELEFONES E EMAILS DA EMPRESA", "Fone3_Empresa",              22, False),
    ("TELEFONES E EMAILS DA EMPRESA", "Email1_Empresa",             28, False),
    ("TELEFONES E EMAILS DA EMPRESA", "Email2_Empresa",             28, False),
    ("TELEFONES E EMAILS DA EMPRESA", "Email3_Empresa",             28, False),
    # CONTATO 1
    ("CONTATO PRINCIPAL",             "Nome_Contato1",              28, True),
    ("CONTATO PRINCIPAL",             "Cargo_Contato1",             20, False),
    ("CONTATO PRINCIPAL",             "Tel_Principal_Contato1",     22, False),
    ("CONTATO PRINCIPAL",             "Tel_WhatsApp_Contato1",      22, False),
    ("CONTATO PRINCIPAL",             "Tel_Comercial_Contato1",     22, False),
    ("CONTATO PRINCIPAL",             "Email_Contato1",             28, False),
    # CONTATO 2 - vai para Anotações do card no CRM
    ("CONTATO 2 (→ Anotações)",       "Nome_Contato2",              28, False),
    ("CONTATO 2 (→ Anotações)",       "Cargo_Contato2",             20, False),
    ("CONTATO 2 (→ Anotações)",       "Tel_Contato2",               22, False),
    ("CONTATO 2 (→ Anotações)",       "Email_Contato2",             28, False),
    # CONTATO 3 - vai para Anotações do card no CRM
    ("CONTATO 3 (→ Anotações)",       "Nome_Contato3",              28, False),
    ("CONTATO 3 (→ Anotações)",       "Cargo_Contato3",             20, False),
    ("CONTATO 3 (→ Anotações)",       "Tel_Contato3",               22, False),
    ("CONTATO 3 (→ Anotações)",       "Email_Contato3",             28, False),
    # CARD
    ("CARD / NEGÓCIO NO CRM",         "SDR_Responsavel",            22, True),
    ("CARD / NEGÓCIO NO CRM",         "Canal_Aquisicao",            22, False),
    ("CARD / NEGÓCIO NO CRM",         "Canal_Aquisicao_Detalhe",   28, False),
    ("CARD / NEGÓCIO NO CRM",         "Tipo_Negocio",               20, False),
    ("CARD / NEGÓCIO NO CRM",         "Valor_Estimado (R$)",        20, False),
    ("CARD / NEGÓCIO NO CRM",         "Data_Prospecao",             18, False),
    ("CARD / NEGÓCIO NO CRM",         "Obs_Gerais_Card",            40, False),
    # ATIVIDADE 1
    ("ATIVIDADES REALIZADAS",         "Ativ1_Data",                 14, False),
    ("ATIVIDADES REALIZADAS",         "Ativ1_Tipo",                 20, False),
    ("ATIVIDADES REALIZADAS",         "Ativ1_Resultado",            50, False),
    # ATIVIDADE 2
    ("ATIVIDADES REALIZADAS",         "Ativ2_Data",                 14, False),
    ("ATIVIDADES REALIZADAS",         "Ativ2_Tipo",                 20, False),
    ("ATIVIDADES REALIZADAS",         "Ativ2_Resultado",            50, False),
    # ATIVIDADE 3
    ("ATIVIDADES REALIZADAS",         "Ativ3_Data",                 14, False),
    ("ATIVIDADES REALIZADAS",         "Ativ3_Tipo",                 20, False),
    ("ATIVIDADES REALIZADAS",         "Ativ3_Resultado",            50, False),
    # ANOTAÇÕES
    ("ANOTAÇÕES",                     "Anotacoes",                  65, False),
]

# --- Instruções (linha de exemplo) ---
instrucoes = [
    "Ex: 12.345.678/0001-99",
    "Ex: EMPRESA ABC LTDA",
    "Ex: Empresa ABC",
    "Ex: empresaabc.com.br",
    "SP",
    "Ex: São Paulo",
    "Ex: Rua X, 123 - Bairro - CEP 00000-000",
    "Ex: 4930202 - Transporte Rodoviário de Carga",
    "Ex: Ate 50 colaboradores",
    "Ex: Ate R$ 10 milhoes",
    "Ex: (11) 3456-7890",
    "Ex: (11) 98765-4321 - WhatsApp",
    "Ex: (11) 3456-7891",
    "Ex: contato@empresa.com.br",
    "Ex: financeiro@empresa.com.br",
    "Ex: vendas@empresa.com.br",
    "Ex: João Silva",
    "Ex: Diretor Comercial",
    "Ex: (11) 98765-4321",
    "Ex: (11) 97654-3210",
    "Ex: (11) 3456-7890 r.12",
    "Ex: joao.silva@empresa.com.br",
    "Ex: Maria Souza  [sera salvo nas Anotacoes do card, nao vinculado]",
    "Ex: Gerente Financeira",
    "Ex: (11) 91234-5678",
    "Ex: maria@empresa.com.br",
    "Ex: Carlos Pereira  [sera salvo nas Anotacoes do card, nao vinculado]",
    "Ex: Socio",
    "Ex: (11) 99876-5432",
    "Ex: carlos@empresa.com.br",
    "Ex: Claudia",
    "Ex: Inbound, Outbound, Indicacao...",
    "Ex: Outbound - Cold call",
    "Nova Venda",
    "Ex: 15000.00",
    "Ex: 05/02/2026",
    "Ex: Empresa interessada em X, aguardando retorno",
    "Ex: 05/02/2026",
    "Ligação",
    "Ex: Atendeu, pediu retorno na quarta-feira",
    "Ex: 07/02/2026",
    "WhatsApp",
    "Ex: Respondeu mensagem, vai avaliar a proposta",
    "Ex: 10/02/2026",
    "Email",
    "Ex: Não respondeu o email, tentarei novamente",
    "Espaço para anotações livres sobre esta empresa ou contato. Use quantas linhas precisar.",
]

# ============================================================
# MONTAGEM DAS LINHAS
# ============================================================

# LINHA 1: Banner de título
ws.row_dimensions[1].height = 30
total_cols = len(columns)
last_col = get_column_letter(total_cols)
ws.merge_cells(f"A1:{last_col}1")
cell = ws["A1"]
cell.value = (
    "PLANILHA DE IMPORTAÇÃO CRM - HSGrowth  |  "
    "Campos obrigatórios marcados com *  |  "
    "Preencha uma empresa por linha  |  "
    "Não altere os nomes das colunas"
)
cell.fill = fill("1A3A5C")
cell.font = Font(bold=True, size=10, color="FFFFFF")
cell.alignment = center()

# LINHA 2: Cabeçalhos de seção (merged por seção)
ws.row_dimensions[2].height = 26
secao_ranges = {}
for col_idx, (secao, col_name, largura, obrig) in enumerate(columns, 1):
    if secao not in secao_ranges:
        secao_ranges[secao] = [col_idx, col_idx]
    else:
        secao_ranges[secao][1] = col_idx

for secao, (start_col, end_col) in secao_ranges.items():
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    if start_col != end_col:
        ws.merge_cells(f"{start_letter}2:{end_letter}2")
    cell = ws[f"{start_letter}2"]
    cell.value = secao
    cell.fill = fill(SECAO_COR_HEADER[secao])
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.alignment = center()

# LINHA 3: Nomes das colunas
ws.row_dimensions[3].height = 42
for col_idx, (secao, col_name, largura, obrig) in enumerate(columns, 1):
    letter = get_column_letter(col_idx)
    cell = ws.cell(row=3, column=col_idx)
    cell.value = f"{col_name} *" if obrig else col_name
    cell.fill = fill(SECAO_COR_BG[secao])
    cell.font = Font(bold=True, size=9, color=SECAO_COR_HEADER[secao])
    cell.alignment = center()
    cell.border = border_all
    ws.column_dimensions[letter].width = largura

# LINHA 4: Instruções
ws.row_dimensions[4].height = 35
for col_idx, instrucao in enumerate(instrucoes, 1):
    cell = ws.cell(row=4, column=col_idx)
    cell.value = instrucao
    cell.fill = fill("F7F7F7")
    cell.font = Font(italic=True, size=8, color="909090")
    cell.alignment = left_wrap()
    cell.border = border_all

# LINHAS DE DADOS: 5 a 1004 (1000 linhas)
for row in range(5, 1005):
    ws.row_dimensions[row].height = 55
    bg = "FAFAFA" if row % 2 == 0 else "FFFFFF"
    for col_idx, (secao, col_name, largura, obrig) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = fill(bg)
        cell.alignment = left_wrap()
        cell.border = border_all
        cell.font = Font(size=9)

# ============================================================
# DROPDOWNS (Data Validation) - error_style="stop" bloqueia entrada inválida
# ============================================================

def make_dv(formula, sqref, error_title, error_msg):
    """Cria DataValidation com bloqueio total de valor inválido."""
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        errorStyle="stop",
        errorTitle=error_title,
        error=error_msg,
        showErrorMessage=True,
    )
    dv.sqref = sqref
    return dv

# Estado (coluna E = 5) — 27 UFs
ws.add_data_validation(make_dv(
    formula="Referência!$E$2:$E$28",
    sqref="E5:E1004",
    error_title="Estado inválido",
    error_msg="Selecione uma UF válida na lista (ex: SP, RJ, MG).",
))

# Faixa de Funcionários (coluna I = 9) — 6 faixas
ws.add_data_validation(make_dv(
    formula="Referência!$A$2:$A$7",
    sqref="I5:I1004",
    error_title="Faixa de funcionários inválida",
    error_msg="Selecione uma das faixas disponíveis na lista (ex: Ate 50 colaboradores).",
))

# Faixa de Faturamento (coluna J = 10) — 6 faixas
ws.add_data_validation(make_dv(
    formula="Referência!$G$2:$G$7",
    sqref="J5:J1004",
    error_title="Faixa de faturamento inválida",
    error_msg="Selecione uma das faixas disponíveis na lista (ex: Ate R$ 10 milhoes).",
))

# Canal de Aquisição (coluna AF = 32) — 6 canais
ws.add_data_validation(make_dv(
    formula="Referência!$B$2:$B$7",
    sqref="AF5:AF1004",
    error_title="Canal inválido",
    error_msg="Selecione: Inbound, Outbound, Indicacao, Parcerias, Eventos ou Base.",
))

# Canal de Aquisição - Detalhe (coluna AG = 33) — 22 opções
ws.add_data_validation(make_dv(
    formula="Referência!$C$2:$C$23",
    sqref="AG5:AG1004",
    error_title="Detalhe de canal inválido",
    error_msg="Selecione um detalhe da lista (ex: Outbound - Cold call, Inbound - SEO).",
))

# Tipo de Negócio (coluna AH = 34) — 3 tipos
ws.add_data_validation(make_dv(
    formula="Referência!$D$2:$D$4",
    sqref="AH5:AH1004",
    error_title="Tipo de negócio inválido",
    error_msg="Selecione: Nova Venda, Cross Sell ou Up Sell.",
))

# Tipo de Atividade 1 (coluna AM = 39) — 8 tipos
ws.add_data_validation(make_dv(
    formula="Referência!$F$2:$F$9",
    sqref="AM5:AM1004",
    error_title="Tipo de atividade inválido",
    error_msg="Selecione: Ligacao, WhatsApp, Email, Reuniao, Visita presencial, LinkedIn, Fale Conosco / Site ou Outro.",
))

# Tipo de Atividade 2 (coluna AP = 42) — 8 tipos
ws.add_data_validation(make_dv(
    formula="Referência!$F$2:$F$9",
    sqref="AP5:AP1004",
    error_title="Tipo de atividade inválido",
    error_msg="Selecione: Ligacao, WhatsApp, Email, Reuniao, Visita presencial, LinkedIn, Fale Conosco / Site ou Outro.",
))

# Tipo de Atividade 3 (coluna AS = 45) — 8 tipos
ws.add_data_validation(make_dv(
    formula="Referência!$F$2:$F$9",
    sqref="AS5:AS1004",
    error_title="Tipo de atividade inválido",
    error_msg="Selecione: Ligacao, WhatsApp, Email, Reuniao, Visita presencial, LinkedIn, Fale Conosco / Site ou Outro.",
))

# ============================================================
# CONGELAR PAINÉIS e ajustes finais
# ============================================================
ws.freeze_panes = "C5"   # congela linhas 1-4 e colunas A-B
ws.sheet_view.zoomScale = 85

# ============================================================
# SALVAR
# ============================================================
output_path = "scripts/Planilha_Importacao_CRM.xlsx"
wb.save(output_path)
print(f"Planilha salva: {output_path}")
print(f"Total de colunas: {len(columns)}")
print(f"Secoes: {[s.encode('ascii', 'replace').decode() for s in secao_ranges.keys()]}")
