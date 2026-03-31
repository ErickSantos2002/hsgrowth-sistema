"""
fix_lealnovo_claudia.py — Corrige a planilha LealNovo - Claudia antes da importação.

Ajustes aplicados:
- SDR_Responsavel *   : preenche "Claudia" em todas as linhas (coluna vazia)
- Faixa de Funcionários : normaliza de "De X a Y" para os valores aceitos pelo CRM
- Faixa de Faturamento  : normaliza de "De R$ X à R$ Y" para os valores aceitos pelo CRM

Obs: Estado (por extenso) já é tratado automaticamente pelo import_from_planilha.py.

Gera: Planilha_Importacao_CRM_LealNovo_Claudia_fixed.xlsx
"""

import os
import sys

import openpyxl

FUNC_MAP = {
    "de 1 a 4":       "Ate 50 colaboradores",
    "de 5 a 9":       "Ate 50 colaboradores",
    "de 10 a 19":     "Ate 50 colaboradores",
    "de 20 a 49":     "Ate 50 colaboradores",
    "de 50 a 99":     "51-100 colaboradores",
    "de 100 a 199":   "101-300 colaboradores",
    "de 200 a 499":   "301-600 colaboradores",
    "de 500 a 1.000": "601-1.000 colaboradores",
    "de 500 a 1000":  "601-1.000 colaboradores",
}

FAT_MAP = {
    "de r$ 900 mil a r$ 4.8 milhoes":     "Ate R$ 10 milhoes",
    "de r$ 4.8 milhoes a r$ 10 milhoes":  "Ate R$ 10 milhoes",
    "de r$ 10 milhoes a r$ 15 milhoes":   "R$ 10-30 milhoes",
    "de r$ 15 milhoes a r$ 20 milhoes":   "R$ 10-30 milhoes",
    "de r$ 20 milhoes a r$ 30 milhoes":   "R$ 10-30 milhoes",
    "de r$ 30 milhoes a r$ 50 milhoes":   "R$ 30-100 milhoes",
    "de r$ 50 milhoes a r$ 75 milhoes":   "R$ 30-100 milhoes",
    "de r$ 75 milhoes a r$ 100 milhoes":  "R$ 30-100 milhoes",
    "de r$ 100 milhoes a r$ 150 milhoes": "R$ 100-300 milhoes",
    "de r$ 150 milhoes a r$ 200 milhoes": "R$ 100-300 milhoes",
    "de r$ 200 milhoes a r$ 300 milhoes": "R$ 100-300 milhoes",
    "de r$ 300 milhoes a r$ 500 milhoes": "R$ 300 milhoes - R$ 1 bilhao",
    "de r$ 500 milhoes a r$ 1 bilhao":    "R$ 300 milhoes - R$ 1 bilhao",
    "acima de r$ 1 bilhao":               "Acima de R$ 1 bilhao",
}

SDR = "Claudia"


def normalize_key(val) -> str:
    if not val:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(val)).encode("ascii", "ignore").decode("ascii")
    return s.lower().strip()


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    src = os.path.join(script_dir, "Planilha Importação LealNovo - Claudia.xlsx")
    dst = os.path.join(script_dir, "Planilha_Importacao_CRM_LealNovo_Claudia_fixed.xlsx")

    if not os.path.exists(src):
        print(f"ERRO: arquivo nao encontrado: {src}")
        sys.exit(1)

    wb = openpyxl.load_workbook(src)
    ws = wb.active

    # Cabeçalho na linha 3, exemplos na linha 4, dados a partir da linha 5
    header_row = 3
    data_start = 5

    headers = {
        ws.cell(row=header_row, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=c).value
    }

    col_sdr  = headers.get("SDR_Responsavel *") or headers.get("SDR_Responsavel")
    col_func = headers.get("Faixa de Funcionários") or headers.get("Faixa_Func")
    col_fat  = headers.get("Faixa de Faturamento") or headers.get("Faixa_Fat")

    print(f"Cabeçalho na linha {header_row}")
    print(f"  col SDR_Responsavel    : {col_sdr}")
    print(f"  col Faixa de Funcionários: {col_func}")
    print(f"  col Faixa de Faturamento : {col_fat}")
    print()

    stats = {"sdr": 0, "func": 0, "fat": 0, "func_miss": [], "fat_miss": []}

    for r in range(data_start, ws.max_row + 1):
        # Pula linhas completamente vazias (verifica col 1 e 2)
        if not ws.cell(row=r, column=1).value and not ws.cell(row=r, column=2).value:
            continue

        # SDR — preenche sempre (estava vazio em todas as linhas)
        if col_sdr:
            ws.cell(row=r, column=col_sdr).value = SDR
            stats["sdr"] += 1

        # Faixa de Funcionários
        if col_func:
            raw = ws.cell(row=r, column=col_func).value
            key = normalize_key(raw)
            mapped = FUNC_MAP.get(key)
            if mapped:
                ws.cell(row=r, column=col_func).value = mapped
                stats["func"] += 1
            elif raw:
                stats["func_miss"].append(f"linha {r}: {raw!r}")

        # Faixa de Faturamento
        if col_fat:
            raw = ws.cell(row=r, column=col_fat).value
            key = normalize_key(raw)
            mapped = FAT_MAP.get(key)
            if mapped:
                ws.cell(row=r, column=col_fat).value = mapped
                stats["fat"] += 1
            elif raw:
                stats["fat_miss"].append(f"linha {r}: {raw!r}")

    wb.save(dst)

    print("=" * 60)
    print("FIX LEALNOVO CLAUDIA — RESULTADO")
    print("=" * 60)
    print(f"  Arquivo gerado       : {os.path.basename(dst)}")
    print(f"  SDR preenchido       : {stats['sdr']} linhas")
    print(f"  Func normalizadas    : {stats['func']}")
    print(f"  Fat  normalizadas    : {stats['fat']}")
    if stats["func_miss"]:
        print(f"  Func sem mapa ({len(stats['func_miss'])}):")
        for m in stats["func_miss"]:
            print(f"    {m}")
    if stats["fat_miss"]:
        print(f"  Fat  sem mapa ({len(stats['fat_miss'])}):")
        for m in stats["fat_miss"]:
            print(f"    {m}")
    print("=" * 60)
    print("\nProximo passo:")
    print(f"  python import_from_planilha.py {os.path.basename(dst)}")


if __name__ == "__main__":
    main()
