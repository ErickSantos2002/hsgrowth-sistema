"""
fix_karolaine.py — Corrige a Planilha_Importacao_CRM_Karolaine.xlsx antes da importação.

Ajustes aplicados:
- Data Criação*          : preenche com hoje onde estiver vazia
- SDR_Responsavel *      : preenche "Karolaine Martins" em todas as linhas
- Canal_Aquisicao        : preenche "Outbound" em todas as linhas
- Canal_Aquisicao_Detalhe: preenche "Outbound - Lista fria" em todas as linhas
- Faixa de Funcionários  : normaliza para os valores aceitos pelo CRM
- Faixa de Faturamento   : normaliza para os valores aceitos pelo CRM

Gera: Planilha_Importacao_CRM_Karolaine_fixed.xlsx
"""

import os
import sys
from datetime import datetime

import openpyxl

FUNC_MAP = {
    "de 1 a 4":       "Ate 50 colaboradores",
    "de 5 a 9":       "Ate 50 colaboradores",
    "de 10 a 19":     "Ate 50 colaboradores",
    "de 20 a 49":     "Ate 50 colaboradores",
    "de 50 a 99":     "51-100 colaboradores",
    "de 100 a 199":   "101-300 colaboradores",
    "de 200 a 499":   "301-600 colaboradores",
    "de 500 a 1.000": "601-1.000 colaboradores",
    "de 500 a 1000":  "601-1.000 colaboradores",
}

FAT_MAP = {
    "de r$ 900 mil a r$ 4.8 milhoes":     "Ate R$ 10 milhoes",
    "de r$ 4.8 milhoes a r$ 10 milhoes":  "Ate R$ 10 milhoes",
    "de r$ 10 milhoes a r$ 15 milhoes":   "R$ 10-30 milhoes",
    "de r$ 15 milhoes a r$ 20 milhoes":   "R$ 10-30 milhoes",
    "de r$ 20 milhoes a r$ 30 milhoes":   "R$ 10-30 milhoes",
    "de r$ 30 milhoes a r$ 50 milhoes":   "R$ 30-100 milhoes",
    "de r$ 50 milhoes a r$ 75 milhoes":   "R$ 30-100 milhoes",
    "de r$ 75 milhoes a r$ 100 milhoes":  "R$ 30-100 milhoes",
    "de r$ 100 milhoes a r$ 150 milhoes": "R$ 100-300 milhoes",
    "de r$ 150 milhoes a r$ 200 milhoes": "R$ 100-300 milhoes",
    "de r$ 200 milhoes a r$ 300 milhoes": "R$ 100-300 milhoes",
    "de r$ 300 milhoes a r$ 500 milhoes": "R$ 300 milhoes - R$ 1 bilhao",
    "de r$ 500 milhoes a r$ 1 bilhao":    "R$ 300 milhoes - R$ 1 bilhao",
    "acima de r$ 1 bilhao":               "Acima de R$ 1 bilhao",
}

SDR       = "Karolaine Martins"
CHANNEL   = "Outbound"
CHANNEL_D = "Outbound - Lista fria"
TODAY     = datetime.now().strftime("%d/%m/%Y")


def normalize_key(val) -> str:
    if not val:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(val)).encode("ascii", "ignore").decode("ascii")
    return s.lower().strip()


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    src = os.path.join(script_dir, "Planilha_Importacao_CRM_Karolaine.xlsx")
    dst = os.path.join(script_dir, "Planilha_Importacao_CRM_Karolaine_fixed.xlsx")

    if not os.path.exists(src):
        print(f"ERRO: arquivo nao encontrado: {src}")
        sys.exit(1)

    wb = openpyxl.load_workbook(src)
    ws = wb.active

    headers = {
        ws.cell(row=3, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=3, column=c).value
    }

    col_data     = headers.get("Data Criação*")
    col_sdr      = headers.get("SDR_Responsavel *")
    col_channel  = headers.get("Canal_Aquisicao")
    col_channel_d= headers.get("Canal_Aquisicao_Detalhe")
    col_func     = headers.get("Faixa de Funcionários")
    col_fat      = headers.get("Faixa de Faturamento")

    stats = {"data": 0, "func": 0, "fat": 0, "func_miss": [], "fat_miss": []}

    for r in range(5, ws.max_row + 1):
        if not ws.cell(row=r, column=2).value and not ws.cell(row=r, column=3).value:
            continue

        # Data de criação → hoje apenas se vazia
        if col_data and not ws.cell(row=r, column=col_data).value:
            ws.cell(row=r, column=col_data).value = TODAY
            stats["data"] += 1

        # SDR
        if col_sdr:
            ws.cell(row=r, column=col_sdr).value = SDR

        # Canal de aquisição
        if col_channel:
            ws.cell(row=r, column=col_channel).value = CHANNEL
        if col_channel_d:
            ws.cell(row=r, column=col_channel_d).value = CHANNEL_D

        # Faixa de Funcionários
        if col_func:
            raw = ws.cell(row=r, column=col_func).value
            key = normalize_key(raw)
            mapped = FUNC_MAP.get(key)
            if mapped:
                ws.cell(row=r, column=col_func).value = mapped
                stats["func"] += 1
            elif raw:
                stats["func_miss"].append(f"linha {r}: {raw!r}")

        # Faixa de Faturamento
        if col_fat:
            raw = ws.cell(row=r, column=col_fat).value
            key = normalize_key(raw)
            mapped = FAT_MAP.get(key)
            if mapped:
                ws.cell(row=r, column=col_fat).value = mapped
                stats["fat"] += 1
            elif raw:
                stats["fat_miss"].append(f"linha {r}: {raw!r}")

    wb.save(dst)

    print("=" * 60)
    print("FIX KAROLAINE — RESULTADO")
    print("=" * 60)
    print(f"  Arquivo gerado     : {os.path.basename(dst)}")
    print(f"  Datas preenchidas  : {stats['data']}")
    print(f"  Func normalizadas  : {stats['func']}")
    print(f"  Fat  normalizadas  : {stats['fat']}")
    if stats["func_miss"]:
        print(f"  Func sem mapa ({len(stats['func_miss'])}):")
        for m in stats["func_miss"]: print(f"    {m}")
    if stats["fat_miss"]:
        print(f"  Fat  sem mapa ({len(stats['fat_miss'])}):")
        for m in stats["fat_miss"]: print(f"    {m}")
    print("=" * 60)
    print("\nProximo passo:")
    print(f"  python import_from_planilha.py {os.path.basename(dst)}")


if __name__ == "__main__":
    main()
