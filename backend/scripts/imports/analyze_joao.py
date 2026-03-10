"""Análise da planilha Planilha_Importacao_CRM_JOAO_VICTOR.xlsx."""
import openpyxl
import sys
import io
from datetime import datetime, date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook('Planilha_Importacao_CRM_JOAO_VICTOR.xlsx')
ws = wb.worksheets[0]

valid_faixa_func = ['Ate 50 colaboradores','51-100 colaboradores','101-300 colaboradores',
                    '301-600 colaboradores','601-1.000 colaboradores','Acima de 1.000 colaboradores']
valid_canal = ['Inbound','Outbound','Indicacao','Parcerias','Eventos','Base']
valid_canal_detalhe = ['Inbound - Conteudo','Inbound - Trafego pago','Inbound - SEO','Inbound - Email marketing',
    'Inbound - Levantada de mao (site / WhatsApp / formulario)','Outbound - Lista fria','Outbound - LinkedIn',
    'Outbound - Cold email','Outbound - Cold call','Indicacao - Cliente','Indicacao - Parceiro',
    'Indicacao - Network pessoal','Parcerias - Co-marketing','Parcerias - Integracao tecnologica',
    'Parcerias - Revenda','Eventos - Feira','Eventos - Webinar','Eventos - Workshop','Eventos - Meetup',
    'Base - Reativacao','Base - Cross-sell','Base - Up-sell']
valid_tipo_negocio = ['Nova Venda','Cross Sell','Up Sell']
valid_estados = ['AC','AL','AM','AP','BA','CE','DF','ES','GO','MA','MG','MS','MT','PA','PB','PE','PI',
                 'PR','RJ','RN','RO','RR','RS','SC','SE','SP','TO']
valid_ativ = ['Ligacao','WhatsApp','Email','Reuniao','Visita presencial','LinkedIn','Fale Conosco / Site','Outro']
valid_faixa_fat = ['Ate R$ 10 milhoes','R$ 10-30 milhoes','R$ 30-100 milhoes','R$ 100-300 milhoes',
                   'R$ 300 milhoes - R$ 1 bilhao','Acima de R$ 1 bilhao']

issues = []
total = 0

for r in range(5, ws.max_row + 1):
    if ws.cell(r, 1).value is None:
        continue
    total += 1
    razao = ws.cell(r, 3).value or f'(linha {r})'
    row_issues = []

    for col, nome in [(3, 'Razao Social'), (18, 'Nome_Contato1'), (32, 'SDR_Responsavel')]:
        val = ws.cell(r, col).value
        if not val or not str(val).strip():
            row_issues.append(f'[OBRIGATORIO] {nome}: vazio')

    cnpj = ws.cell(r, 2).value
    if cnpj:
        d = str(cnpj).strip().replace('.', '').replace('/', '').replace('-', '').replace(' ', '')
        if len(d) != 14 or not d.isdigit():
            row_issues.append(f'[FORMATO] CNPJ: "{cnpj}"')
    else:
        row_issues.append('[OBRIGATORIO] CNPJ: vazio')

    for val, valids, nome in [
        (ws.cell(r, 6).value,  valid_estados,      'Estado'),
        (ws.cell(r, 10).value, valid_faixa_func,    'Faixa_Funcionarios'),
        (ws.cell(r, 11).value, valid_faixa_fat,     'Faixa_Faturamento'),
        (ws.cell(r, 33).value, valid_canal,          'Canal_Aquisicao'),
        (ws.cell(r, 34).value, valid_canal_detalhe,  'Canal_Aquisicao_Detalhe'),
        (ws.cell(r, 35).value, valid_tipo_negocio,   'Tipo_Negocio'),
    ]:
        if val and str(val).strip() not in valids:
            row_issues.append(f'[LISTA] {nome}: "{val}"')

    for col, nome in [(40, 'Ativ1_Tipo'), (43, 'Ativ2_Tipo'), (46, 'Ativ3_Tipo')]:
        v = ws.cell(r, col).value
        if v and str(v).strip() not in valid_ativ:
            row_issues.append(f'[LISTA] {nome}: "{v}"')

    for col, nome in [(1, 'Data_Criacao'), (37, 'Data_Prospecao'), (39, 'Ativ1_Data'),
                      (42, 'Ativ2_Data'), (45, 'Ativ3_Data')]:
        v = ws.cell(r, col).value
        if v is not None and not isinstance(v, (datetime, date)):
            row_issues.append(f'[FORMATO] {nome}: "{v}" (texto)')

    valor = ws.cell(r, 36).value
    if valor is not None:
        try:
            float(str(valor).replace(',', '.'))
        except ValueError:
            row_issues.append(f'[FORMATO] Valor_Estimado: "{valor}"')

    if row_issues:
        issues.append((r, str(razao), row_issues))

sep = '=' * 65
print(sep)
print('ANALISE - PLANILHA JOAO VICTOR')
print(sep)
print(f'Linhas analisadas  : {total}')
print(f'Linhas com problema: {len(issues)}')
print(f'Linhas OK          : {total - len(issues)}')
print()
if issues:
    for row_num, empresa, probs in issues:
        print(f'Linha {row_num} | {empresa}:')
        for p in probs:
            print(f'    {p}')
        print()
else:
    print('Nenhum problema encontrado! Planilha pronta para importacao.')
print(sep)
