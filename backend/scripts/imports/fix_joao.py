"""
Script de correção automática da planilha Planilha_Importacao_CRM_JOAO_VICTOR.xlsx.

Correções aplicadas:
  1. SDR_Responsavel: preenchido com "João Victor" em todas as linhas
  2. Nome_Contato1: preenchido com o primeiro nome da Razão Social
  3. Tel_Principal_Contato1: preenchido com Fone1_Empresa (quando disponível)
  4. CNPJ: zero à esquerda restaurado (zfill 14 dígitos, formata XX.XXX.XXX/XXXX-XX)
  5. Estado (UF): nome completo → sigla
  6. Faixa de Faturamento: faixas específicas → padrão do sistema

Gera um novo arquivo _fixed.xlsx sem modificar o original.
"""
import openpyxl
import sys
import io
import unicodedata

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

INPUT_FILE  = 'Planilha_Importacao_CRM_JOAO_VICTOR.xlsx'
OUTPUT_FILE = 'Planilha_Importacao_CRM_JOAO_VICTOR_fixed.xlsx'

# ============================================================
# Mapeamentos
# ============================================================

ESTADO_MAP = {
    'ACRE': 'AC', 'ALAGOAS': 'AL', 'AMAPA': 'AP', 'AMAZONAS': 'AM',
    'BAHIA': 'BA', 'CEARA': 'CE', 'DISTRITO FEDERAL': 'DF', 'ESPIRITO SANTO': 'ES',
    'GOIAS': 'GO', 'MARANHAO': 'MA', 'MATO GROSSO DO SUL': 'MS', 'MATO GROSSO': 'MT',
    'MINAS GERAIS': 'MG', 'PARA': 'PA', 'PARAIBA': 'PB', 'PARANA': 'PR',
    'PERNAMBUCO': 'PE', 'PIAUI': 'PI', 'RIO DE JANEIRO': 'RJ',
    'RIO GRANDE DO NORTE': 'RN', 'RIO GRANDE DO SUL': 'RS', 'RONDONIA': 'RO',
    'RORAIMA': 'RR', 'SANTA CATARINA': 'SC', 'SAO PAULO': 'SP',
    'SERGIPE': 'SE', 'TOCANTINS': 'TO',
}

FATURAMENTO_MAP = {
    'De R$ 900 mil à R$ 4.8 milhões':     'Ate R$ 10 milhoes',
    'De R$ 4.8 milhões à R$ 10 milhões':  'Ate R$ 10 milhoes',
    'De R$ 10 milhões à R$ 15 milhões':   'R$ 10-30 milhoes',
    'De R$ 15 milhões à R$ 20 milhões':   'R$ 10-30 milhoes',
    'De R$ 20 milhões à R$ 30 milhões':   'R$ 10-30 milhoes',
    'De R$ 30 milhões à R$ 50 milhões':   'R$ 30-100 milhoes',
    'De R$ 50 milhões à R$ 75 milhões':   'R$ 30-100 milhoes',
    'De R$ 75 milhões à R$ 100 milhões':  'R$ 30-100 milhoes',
    'De R$ 100 milhões à R$ 150 milhões': 'R$ 100-300 milhoes',
    'De R$ 150 milhões à R$ 200 milhões': 'R$ 100-300 milhoes',
    'De R$ 200 milhões à R$ 300 milhões': 'R$ 100-300 milhoes',
    'De R$ 300 milhões à R$ 500 milhões': 'R$ 300 milhoes - R$ 1 bilhao',
    'De R$ 500 milhões à R$ 1 bilhão':    'R$ 300 milhoes - R$ 1 bilhao',
    'Acima de R$ 1 bilhão':               'Acima de R$ 1 bilhao',
    'Acima de R$ 2 bilhões':              'Acima de R$ 1 bilhao',
}

VALID_ESTADOS   = set(ESTADO_MAP.values())
VALID_FAIXA_FAT = set(FATURAMENTO_MAP.values())

# Índices das colunas (1-indexed, linha 3 = headers)
COL_CNPJ          = 2
COL_RAZAO_SOCIAL  = 3
COL_ESTADO        = 6
COL_FONE1_EMPRESA = 12
COL_FAIXA_FAT     = 11
COL_NOME_CONTATO1 = 18
COL_TEL_CONTATO1  = 20   # Tel_Principal_Contato1
COL_SDR           = 32


def remove_accents(text: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFD', text)
        if unicodedata.category(c) != 'Mn'
    )


def fix_cnpj(value) -> str | None:
    """
    Restaura zero à esquerda perdido pelo Excel e formata como XX.XXX.XXX/XXXX-XX.
    Ex: 4382714000162 (13 dígitos) → 04.382.714/0001-62
    """
    if value is None:
        return None
    digits = ''.join(c for c in str(value).strip() if c.isdigit())
    if not digits:
        return None
    # Restaura zero à esquerda se ficou com 13 dígitos
    digits = digits.zfill(14)
    if len(digits) != 14:
        return str(value)  # não conseguiu corrigir, mantém original
    # Formata: XX.XXX.XXX/XXXX-XX
    return f'{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}'


def fix_estado(value) -> str | None:
    if value is None:
        return value
    v = remove_accents(str(value).strip().upper())
    if v in VALID_ESTADOS:
        return v
    return ESTADO_MAP.get(v, value)


def fix_faturamento(value) -> str | None:
    if value is None:
        return value
    v = str(value).strip()
    if v in VALID_FAIXA_FAT:
        return v
    return FATURAMENTO_MAP.get(v, value)


def first_word_of_company(razao_social: str) -> str:
    """
    Retorna o primeiro nome relevante da Razão Social como nome de contato.
    Ignora sufixos corporativos comuns (LTDA, S/A, etc.) se forem a primeira palavra.
    Ex: 'TRANSITA TRANSPORTES LTDA' → 'Transita'
    """
    stopwords = {'LTDA', 'SA', 'S/A', 'S.A', 'S.A.', 'EIRELI', 'ME', 'EPP', 'CIA', 'CIA.'}
    words = razao_social.strip().split()
    for word in words:
        clean = word.strip('.,/-').upper()
        if clean not in stopwords and len(clean) > 1:
            # Capitaliza: primeira letra maiúscula, resto minúsculo
            return word.strip('.,/-').capitalize()
    # Fallback: primeira palavra mesmo assim
    return words[0].capitalize() if words else razao_social


# ============================================================
# Execução
# ============================================================

wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.worksheets[0]

corrections = {
    'sdr': 0,
    'nome_contato1': 0,
    'tel_contato1': 0,
    'cnpj': 0,
    'estado': 0,
    'faixa_fat': 0,
}
total_rows = 0

for r in range(5, ws.max_row + 1):
    if ws.cell(r, 1).value is None:
        continue
    total_rows += 1

    # --- SDR_Responsavel ---
    cell_sdr = ws.cell(r, COL_SDR)
    if not cell_sdr.value or not str(cell_sdr.value).strip():
        cell_sdr.value = 'João Victor'
        corrections['sdr'] += 1

    # --- Nome_Contato1 ---
    cell_nome = ws.cell(r, COL_NOME_CONTATO1)
    if not cell_nome.value or not str(cell_nome.value).strip():
        razao = ws.cell(r, COL_RAZAO_SOCIAL).value
        if razao:
            cell_nome.value = first_word_of_company(str(razao))
            corrections['nome_contato1'] += 1

    # --- Tel_Principal_Contato1 (usa Fone1_Empresa como fallback) ---
    cell_tel = ws.cell(r, COL_TEL_CONTATO1)
    if not cell_tel.value or not str(cell_tel.value).strip():
        fone1 = ws.cell(r, COL_FONE1_EMPRESA).value
        if fone1:
            cell_tel.value = str(fone1).strip()
            corrections['tel_contato1'] += 1

    # --- CNPJ ---
    cell_cnpj = ws.cell(r, COL_CNPJ)
    original_cnpj = cell_cnpj.value
    fixed_cnpj = fix_cnpj(original_cnpj)
    if fixed_cnpj and fixed_cnpj != str(original_cnpj or '').strip():
        cell_cnpj.value = fixed_cnpj
        # Garante que o Excel trate como texto (evita remover zeros novamente)
        cell_cnpj.number_format = '@'
        corrections['cnpj'] += 1

    # --- Estado (UF) ---
    cell_estado = ws.cell(r, COL_ESTADO)
    original_estado = cell_estado.value
    fixed_estado = fix_estado(original_estado)
    if fixed_estado != original_estado:
        cell_estado.value = fixed_estado
        corrections['estado'] += 1

    # --- Faixa de Faturamento ---
    cell_fat = ws.cell(r, COL_FAIXA_FAT)
    original_fat = cell_fat.value
    fixed_fat = fix_faturamento(original_fat)
    if fixed_fat != original_fat:
        cell_fat.value = fixed_fat
        corrections['faixa_fat'] += 1

wb.save(OUTPUT_FILE)

# ============================================================
# Relatório
# ============================================================
sep = '=' * 60
print(sep)
print('CORRECAO AUTOMATICA - PLANILHA JOAO VICTOR')
print(sep)
print(f'Arquivo gerado   : {OUTPUT_FILE}')
print(f'Linhas corrigidas: {total_rows}/{total_rows}')
print()
print('Correcoes por campo:')
print(f'  SDR_Responsavel       : {corrections["sdr"]} celulas → "João Victor"')
print(f'  Nome_Contato1         : {corrections["nome_contato1"]} celulas → 1º nome da empresa')
print(f'  Tel_Principal_Contato1: {corrections["tel_contato1"]} celulas → Fone1_Empresa')
print(f'  CNPJ                  : {corrections["cnpj"]} celulas → zero restaurado + formatado')
print(f'  Estado (UF)           : {corrections["estado"]} celulas → sigla')
print(f'  Faixa de Faturamento  : {corrections["faixa_fat"]} celulas → padrao do sistema')
print(sep)
