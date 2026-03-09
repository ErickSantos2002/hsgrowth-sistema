"""
Script de análise da planilha Planilha_Importacao_CRM_CLAUDIA_SILVA.xlsx.
Verifica campos obrigatórios, valores fora do padrão e formatos inválidos.
"""
import openpyxl
import sys
import io
from datetime import datetime, date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook('Planilha_Importacao_CRM_CLAUDIA_SILVA.xlsx')
ws = wb.worksheets[0]

# Valores válidos conforme aba Referência
valid_faixa_func = [
    'Ate 50 colaboradores',
    '51-100 colaboradores',
    '101-300 colaboradores',
    '301-600 colaboradores',
    '601-1.000 colaboradores',
    'Acima de 1.000 colaboradores',
]
valid_canal = ['Inbound', 'Outbound', 'Indicacao', 'Parcerias', 'Eventos', 'Base']
valid_canal_detalhe = [
    'Inbound - Conteudo',
    'Inbound - Trafego pago',
    'Inbound - SEO',
    'Inbound - Email marketing',
    'Inbound - Levantada de mao (site / WhatsApp / formulario)',
    'Outbound - Lista fria',
    'Outbound - LinkedIn',
    'Outbound - Cold email',
    'Outbound - Cold call',
    'Indicacao - Cliente',
    'Indicacao - Parceiro',
    'Indicacao - Network pessoal',
    'Parcerias - Co-marketing',
    'Parcerias - Integracao tecnologica',
    'Parcerias - Revenda',
    'Eventos - Feira',
    'Eventos - Webinar',
    'Eventos - Workshop',
    'Eventos - Meetup',
    'Base - Reativacao',
    'Base - Cross-sell',
    'Base - Up-sell',
]
valid_tipo_negocio = ['Nova Venda', 'Cross Sell', 'Up Sell']
valid_estados = [
    'AC', 'AL', 'AM', 'AP', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA',
    'MG', 'MS', 'MT', 'PA', 'PB', 'PE', 'PI', 'PR', 'RJ', 'RN',
    'RO', 'RR', 'RS', 'SC', 'SE', 'SP', 'TO',
]
valid_ativ = [
    'Ligacao', 'WhatsApp', 'Email', 'Reuniao',
    'Visita presencial', 'LinkedIn', 'Fale Conosco / Site', 'Outro',
]
valid_faixa_fat = [
    'Ate R$ 10 milhoes',
    'R$ 10-30 milhoes',
    'R$ 30-100 milhoes',
    'R$ 100-300 milhoes',
    'R$ 300 milhoes - R$ 1 bilhao',
    'Acima de R$ 1 bilhao',
]

# Mapeamento de colunas da planilha CLAUDIA (linha 3 = headers, dados a partir da linha 5)
# Col 1: Data Criação*, Col 2: CNPJ, Col 3: Razão Social, Col 4: Nome Fantasia,
# Col 5: Site, Col 6: Estado (UF), Col 7: Cidade, Col 8: Endereço Completo,
# Col 9: CNAE Principal, Col 10: Faixa de Funcionários, Col 11: Faixa de Faturamento,
# Col 12-14: Fones Empresa, Col 15-17: Emails Empresa,
# Col 18: Nome_Contato1, Col 19: Cargo, Col 20: Tel_Principal, Col 21: Tel_WhatsApp,
# Col 22: Tel_Comercial, Col 23: Email_Contato1,
# Col 24-27: Contato2, Col 28-31: Contato3,
# Col 32: SDR_Responsavel, Col 33: Canal_Aquisicao, Col 34: Canal_Aquisicao_Detalhe,
# Col 35: Tipo_Negocio, Col 36: Valor_Estimado, Col 37: Data_Prospecao, Col 38: Obs_Gerais,
# Col 39-41: Ativ1, Col 42-44: Ativ2, Col 45-47: Ativ3, Col 48: Anotacoes

issues = []
total_rows = 0

for r in range(5, ws.max_row + 1):
    # Só processa linhas com col A preenchida
    if ws.cell(r, 1).value is None:
        continue

    total_rows += 1
    razao_social = ws.cell(r, 3).value or f'(linha {r} sem razao social)'
    row_issues = []

    # --- Campos obrigatórios ---
    for col, nome in [(3, 'Razao Social'), (18, 'Nome_Contato1'), (32, 'SDR_Responsavel')]:
        val = ws.cell(r, col).value
        if not val or not str(val).strip():
            row_issues.append(f'[OBRIGATORIO] {nome}: campo vazio')

    # --- CNPJ: 14 dígitos numéricos ---
    cnpj = ws.cell(r, 2).value
    if cnpj:
        cnpj_str = (
            str(cnpj).strip()
            .replace('.', '').replace('/', '').replace('-', '').replace(' ', '')
        )
        if len(cnpj_str) != 14 or not cnpj_str.isdigit():
            row_issues.append(f'[FORMATO] CNPJ: "{cnpj}" invalido (esperado XX.XXX.XXX/XXXX-XX)')
    else:
        row_issues.append('[OBRIGATORIO] CNPJ: campo vazio')

    # --- Estado (UF) ---
    estado = ws.cell(r, 6).value
    if estado and str(estado).strip() not in valid_estados:
        row_issues.append(f'[LISTA] Estado: "{estado}" nao reconhecido')

    # --- Faixa de Funcionários ---
    ff = ws.cell(r, 10).value
    if ff and str(ff).strip() not in valid_faixa_func:
        row_issues.append(f'[LISTA] Faixa_Funcionarios: "{ff}" fora do padrao')

    # --- Faixa de Faturamento ---
    fat = ws.cell(r, 11).value
    if fat and str(fat).strip() not in valid_faixa_fat:
        row_issues.append(f'[LISTA] Faixa_Faturamento: "{fat}" fora do padrao')

    # --- Canal Aquisicao ---
    canal = ws.cell(r, 33).value
    if canal and str(canal).strip() not in valid_canal:
        row_issues.append(f'[LISTA] Canal_Aquisicao: "{canal}" fora do padrao')

    # --- Canal Aquisicao Detalhe ---
    canal_det = ws.cell(r, 34).value
    if canal_det and str(canal_det).strip() not in valid_canal_detalhe:
        row_issues.append(f'[LISTA] Canal_Aquisicao_Detalhe: "{canal_det}" fora do padrao')

    # --- Tipo Negocio ---
    tipo_neg = ws.cell(r, 35).value
    if tipo_neg and str(tipo_neg).strip() not in valid_tipo_negocio:
        row_issues.append(f'[LISTA] Tipo_Negocio: "{tipo_neg}" fora do padrao')

    # --- Tipos de Atividade ---
    for col, nome in [(40, 'Ativ1_Tipo'), (43, 'Ativ2_Tipo'), (46, 'Ativ3_Tipo')]:
        v = ws.cell(r, col).value
        if v and str(v).strip() not in valid_ativ:
            row_issues.append(f'[LISTA] {nome}: "{v}" fora do padrao')

    # --- Datas: devem ser objetos date/datetime, não texto ---
    for col, nome in [
        (1, 'Data_Criacao'),
        (37, 'Data_Prospecao'),
        (39, 'Ativ1_Data'),
        (42, 'Ativ2_Data'),
        (45, 'Ativ3_Data'),
    ]:
        v = ws.cell(r, col).value
        if v is not None and not isinstance(v, (datetime, date)):
            row_issues.append(f'[FORMATO] {nome}: "{v}" nao e data (texto puro)')

    # --- Valor Estimado: deve ser número ---
    valor = ws.cell(r, 36).value
    if valor is not None:
        try:
            float(str(valor).replace(',', '.'))
        except ValueError:
            row_issues.append(f'[FORMATO] Valor_Estimado: "{valor}" nao e um numero valido')

    if row_issues:
        issues.append((r, str(razao_social), row_issues))

# --- Relatório final ---
sep = '=' * 70
sep2 = '-' * 70

print(sep)
print('ANALISE - PLANILHA CLAUDIA SILVA')
print(sep)
print(f'Total de linhas analisadas : {total_rows}')
print(f'Linhas com problemas       : {len(issues)}')
print(f'Linhas OK (sem problemas)  : {total_rows - len(issues)}')
print()

if issues:
    print(sep2)
    print('DETALHAMENTO DOS PROBLEMAS:')
    print(sep2)
    for row_num, empresa, probs in issues:
        print(f'\nLinha {row_num} | {empresa}:')
        for p in probs:
            print(f'    {p}')
else:
    print('Nenhum problema encontrado! Planilha dentro do padrao.')

print()
print(sep)
