"""
Script 1 de 2 — Normalização de dados da Planilha_Importacao_CRM_Novos_SDR.xlsx

O que faz:
  - Converte Faixa de Funcionários para os valores aceitos pelo CRM
  - Converte Faixa de Faturamento para os valores aceitos pelo CRM
  - Processa TODAS as 2189 linhas de dados
  - Salva o resultado em Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx

Rode ANTES do fix_novos_sdr_import.py
"""

import os
import openpyxl
from unicodedata import normalize as unicode_normalize

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE  = os.path.join(SCRIPT_DIR, "Planilha_Importacao_CRM_Novos_SDR.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx")

HEADER_ROW = 3   # linha com os nomes das colunas
DATA_START  = 4  # primeira linha de dados reais

# ── Mapeamentos ────────────────────────────────────────────────────────────────

FAIXA_FUNC_MAP = {
    "de 1 a 4":    "Ate 50 colaboradores",
    "de 5 a 9":    "Ate 50 colaboradores",
    "de 10 a 19":  "Ate 50 colaboradores",
    "de 20 a 49":  "Ate 50 colaboradores",
    "de 50 a 99":  "51-100 colaboradores",
    "de 100 a 199": "101-300 colaboradores",
    "de 200 a 499": "301-600 colaboradores",
    "de 500 a 999": "601-1.000 colaboradores",
    "de 500 a 1.000": "601-1.000 colaboradores",
    "acima de 1000": "Acima de 1.000 colaboradores",
    "acima de 1.000": "Acima de 1.000 colaboradores",
}

FAIXA_FAT_MAP = {
    "de r$ 900 mil a r$ 4.8 milhoes":   "Ate R$ 10 milhoes",
    "de r$ 4.8 milhoes a r$ 10 milhoes": "Ate R$ 10 milhoes",
    "de r$ 10 milhoes a r$ 15 milhoes":  "R$ 10-30 milhoes",
    "de r$ 15 milhoes a r$ 20 milhoes":  "R$ 10-30 milhoes",
    "de r$ 20 milhoes a r$ 30 milhoes":  "R$ 10-30 milhoes",
    "de r$ 30 milhoes a r$ 50 milhoes":  "R$ 30-100 milhoes",
    "de r$ 50 milhoes a r$ 75 milhoes":  "R$ 30-100 milhoes",
    "de r$ 75 milhoes a r$ 100 milhoes": "R$ 30-100 milhoes",
    "de r$ 100 milhoes a r$ 150 milhoes": "R$ 100-300 milhoes",
    "de r$ 150 milhoes a r$ 200 milhoes": "R$ 100-300 milhoes",
    "de r$ 200 milhoes a r$ 300 milhoes": "R$ 100-300 milhoes",
    "de r$ 300 milhoes a r$ 500 milhoes": "R$ 300 milhoes - R$ 1 bilhao",
    "de r$ 500 milhoes a r$ 1 bilhao":    "R$ 300 milhoes - R$ 1 bilhao",
    "de r$ 1 bilhao a r$ 2 bilhoes":      "Acima de R$ 1 bilhao",
    "acima de r$ 1 bilhao":  "Acima de R$ 1 bilhao",
    "acima de r$ 2 bilhoes": "Acima de R$ 1 bilhao",
}


def strip_accents(text: str) -> str:
    return unicode_normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def normalize_key(val) -> str:
    if val is None:
        return ""
    return strip_accents(str(val).strip().lower())


def find_col(ws, header_row: int, name: str) -> int | None:
    name_norm = strip_accents(name.lower().strip())
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col).value
        if cell_val and strip_accents(str(cell_val).lower().strip()) == name_norm:
            return col
    return None


def main():
    print("=" * 60)
    print("SCRIPT 1 — Normalização de dados (todas as linhas)")
    print("=" * 60)
    print(f"Entrada : {INPUT_FILE}")
    print(f"Saída   : {OUTPUT_FILE}")

    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb["Importação_CRM"]

    # Localiza colunas pelo nome do cabeçalho
    col_func = find_col(ws, HEADER_ROW, "Faixa de Funcionários")
    col_fat  = find_col(ws, HEADER_ROW, "Faixa de Faturamento")

    if not col_func:
        print("ERRO: Coluna 'Faixa de Funcionários' não encontrada.")
        return
    if not col_fat:
        print("ERRO: Coluna 'Faixa de Faturamento' não encontrada.")
        return

    print(f"\nColuna Faixa Funcionários : col {col_func}")
    print(f"Coluna Faixa Faturamento  : col {col_fat}")
    print(f"\nProcessando linhas {DATA_START} até {ws.max_row}...\n")

    stats = {"func_ok": 0, "func_null": 0, "fat_ok": 0, "fat_null": 0, "rows": 0}
    unmapped_func = set()
    unmapped_fat  = set()

    for row in range(DATA_START, ws.max_row + 1):
        # Pula linhas completamente vazias
        if not ws.cell(row, 1).value:
            continue

        stats["rows"] += 1

        # ── Faixa de Funcionários ─────────────────────────────────
        raw_func = ws.cell(row, col_func).value
        key_func = normalize_key(raw_func)
        if key_func:
            mapped = FAIXA_FUNC_MAP.get(key_func)
            if mapped:
                ws.cell(row, col_func).value = mapped
                stats["func_ok"] += 1
            else:
                ws.cell(row, col_func).value = None
                stats["func_null"] += 1
                unmapped_func.add(str(raw_func))

        # ── Faixa de Faturamento ──────────────────────────────────
        raw_fat = ws.cell(row, col_fat).value
        key_fat = normalize_key(raw_fat)
        if key_fat:
            mapped = FAIXA_FAT_MAP.get(key_fat)
            if mapped:
                ws.cell(row, col_fat).value = mapped
                stats["fat_ok"] += 1
            else:
                ws.cell(row, col_fat).value = None
                stats["fat_null"] += 1
                unmapped_fat.add(str(raw_fat))

    wb.save(OUTPUT_FILE)

    print("=" * 60)
    print("RELATÓRIO")
    print("=" * 60)
    print(f"  Linhas processadas         : {stats['rows']}")
    print(f"  Faixa Func normalizadas    : {stats['func_ok']}")
    print(f"  Faixa Func zeradas (null)  : {stats['func_null']}")
    print(f"  Faixa Fat normalizadas     : {stats['fat_ok']}")
    print(f"  Faixa Fat zeradas (null)   : {stats['fat_null']}")
    if unmapped_func:
        print(f"\n  Valores Func sem mapeamento: {sorted(unmapped_func)}")
    if unmapped_fat:
        print(f"\n  Valores Fat sem mapeamento : {sorted(unmapped_fat)}")
    print(f"\nArquivo salvo em: {OUTPUT_FILE}")
    print("=" * 60)
    print("Próximo passo: rode fix_novos_sdr_import.py")


if __name__ == "__main__":
    main()
