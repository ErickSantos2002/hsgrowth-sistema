"""
Padroniza a Planilha_Nova_030726.xlsx (prospecção de transportadoras/operadores
logísticos) para o layout aceito pelo import_from_planilha.py.

Entrada : Planilha_Nova_030726.xlsx  (aba CONSOLIDADO, cabeçalho na row 3, dados row 4+)
Saída   : Planilha_Transportadoras_Prospeccao_fixed.xlsx  (aba "Importação_CRM",
          3 linhas de cabeçalho copiadas do template padrão, dados a partir da row 4,
          coluna Status_Importacao vazia = pendente)

Regras de conversão:
  - Nº FUNC. (número) -> Faixa de Funcionários (texto). 0/vazio => em branco.
  - FATURAMENTO (R$)  -> Faixa de Faturamento (texto). 0/vazio => em branco.
  - Site: usa o hyperlink real da coluna WEBSITE. "🔍 Buscar site" (busca Google) => em branco.
  - LinkedIn da empresa: hyperlink da coluna "LinkedIn Empresa" -> nova coluna
    "Linkedin_Empresa" (col 52), lida pelo importer em client.linkedin_url.
    URLs genéricas (.../company/ sem nome) => em branco.
  - SSMA no LinkedIn: hyperlink (busca no Google) -> nova coluna "Notes_Cliente"
    (col 53), lida pelo importer em client.notes ("Observações" do cliente).
    Fica visível na seção "Cliente (Organização)" do detalhe do card.
  - Nome_Contato1 fica em branco (planilha nova não tem contato dedicado).
  - SDR_Responsavel / Canal_Aquisicao / Status_Importacao ficam em branco
    (preenchidos pelos scripts de lote).

IDEMPOTÊNCIA: se o arquivo de saída já existir, o controle de importação
('Importado' + SDR/Canal por CNPJ) é preservado — assim regerar NÃO re-importa
lotes já subidos (evita duplicatas).

ANTI-DUPLICATA (CNPJ já no CRM): a cada execução consulta os CNPJs já cadastrados
no CRM. As linhas pendentes cujo CNPJ já existe são jogadas para o FINAL e marcadas
Status_Importacao = "CNPJ_no_CRM" — os geradores de lote pulam qualquer linha com
Status preenchido, então elas não sobem (ficam para revisão futura). Rode a partir
da pasta backend/ (precisa do .env para conectar no banco).
"""

import os
import openpyxl

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
SRC_FILE     = os.path.join(SCRIPT_DIR, "Planilha_Nova_030726.xlsx")
TEMPLATE_STD = os.path.join(SCRIPT_DIR, "Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx")
OUT_FILE     = os.path.join(SCRIPT_DIR, "Planilha_Transportadoras_Prospeccao_fixed.xlsx")

SRC_SHEET    = "CONSOLIDADO"
SRC_DATA     = 4
STD_SHEET    = "Importação_CRM"
STD_HEADER   = 3      # 3 linhas de cabeçalho no padrão
STD_COLS     = 51

# Índices (1-based) das colunas na planilha NOVA (aba CONSOLIDADO)
N_LOTE, N_CNPJ, N_RAZAO, N_FANTASIA, N_CIDADE, N_UF = 1, 2, 3, 4, 5, 6
N_PORTE, N_FUNC, N_FATUR, N_CNAE, N_TEL, N_EMAIL1, N_EMAIL2 = 7, 8, 9, 10, 11, 12, 13
N_WEBSITE, N_SSMA, N_LINKEDIN = 14, 16, 17

# Índices (1-based) das colunas no PADRÃO (aba Importação_CRM)
P_CNPJ, P_RAZAO, P_FANTASIA, P_SITE, P_UF, P_CIDADE, P_ENDERECO = 1, 2, 3, 4, 5, 6, 7
P_CNAE, P_FX_FUNC, P_FX_FATUR = 8, 9, 10
P_FONE1, P_EMAIL1, P_EMAIL2 = 11, 14, 15
P_SDR, P_CANAL, P_DETALHE = 34, 35, 36
P_STATUS = 51
P_LINKEDIN_EMP = 52   # NOVA coluna: LinkedIn da empresa -> client.linkedin_url
P_NOTES_CLI = 53      # NOVA coluna: Observações do cliente -> client.notes (link SSMA)
P_VENDEDOR = 54       # NOVA coluna: Vendedor_Responsavel -> card.assigned_to_id
                      # (o template padrão só tem SDR_Responsavel; esta coluna permite
                      #  atribuir vendedores. Preenchida pelos scripts de lote.)

STATUS_IMPORTED = "Importado"
STATUS_DUP_CRM  = "CNPJ_no_CRM"   # marca linhas cujo CNPJ já está cadastrado no CRM


def cnpj14(v):
    """Normaliza qualquer CNPJ (formatado ou não) para 14 dígitos, ou None."""
    if v is None:
        return None
    d = "".join(c for c in str(v) if c.isdigit())
    return d.zfill(14) if d else None


def load_crm_cnpjs():
    """Consulta o CRM e retorna o conjunto de CNPJs (14 dígitos) já cadastrados."""
    import sys
    backend_dir = os.path.dirname(os.path.dirname(SCRIPT_DIR))  # .../backend
    if backend_dir not in sys.path:
        sys.path.insert(0, backend_dir)
    from app.db.session import SessionLocal
    from app.models.client import Client

    db = SessionLocal()
    try:
        docs = db.query(Client.document).filter(Client.is_deleted == False).all()
    finally:
        db.close()
    return {c for (doc,) in docs if (c := cnpj14(doc))}


def faixa_funcionarios(val):
    try:
        n = int(float(val))
    except (TypeError, ValueError):
        return None
    if n <= 0:
        return None
    if n <= 50:
        return "Ate 50 colaboradores"
    if n <= 100:
        return "51-100 colaboradores"
    if n <= 300:
        return "101-300 colaboradores"
    if n <= 600:
        return "301-600 colaboradores"
    if n <= 1000:
        return "601-1.000 colaboradores"
    return "Acima de 1.000 colaboradores"


def faixa_faturamento(val):
    try:
        n = float(val)
    except (TypeError, ValueError):
        return None
    if n <= 0:
        return None
    if n <= 10_000_000:
        return "Ate R$ 10 milhoes"
    if n <= 30_000_000:
        return "R$ 10-30 milhoes"
    if n <= 100_000_000:
        return "R$ 30-100 milhoes"
    if n <= 300_000_000:
        return "R$ 100-300 milhoes"
    if n <= 1_000_000_000:
        return "R$ 300 milhoes - R$ 1 bilhao"
    return "Acima de R$ 1 bilhao"


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def cell_link(ws, r, c):
    cell = ws.cell(r, c)
    return cell.hyperlink.target if cell.hyperlink else None


def site_url(ws, r):
    """URL real do site; None quando é placeholder de busca."""
    val = clean(ws.cell(r, N_WEBSITE).value)
    if not val or "buscar" in val.lower():
        return None
    return cell_link(ws, r, N_WEBSITE) or val


def linkedin_empresa(ws, r):
    """LinkedIn da empresa; descarta stub genérico .../company/ sem nome."""
    link = cell_link(ws, r, N_LINKEDIN)
    if not link:
        return None
    if "/company/" in link:
        slug = link.split("/company/", 1)[1].strip("/ ").strip()
        if not slug:
            return None
    return link


def ssma_obs(ws, r):
    """Texto de observação com o link de busca do SSMA no LinkedIn."""
    link = cell_link(ws, r, N_SSMA)
    if not link:
        return None
    return f"🔍 Buscar SSMA no LinkedIn: {link}"


def load_prior_control():
    """Carrega SDR/Canal/Detalhe/Status do arquivo de saída anterior (por CNPJ)."""
    prior = {}
    if not os.path.exists(OUT_FILE):
        return prior
    wb = openpyxl.load_workbook(OUT_FILE, read_only=True)
    ws = wb[STD_SHEET]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= STD_HEADER:
            continue
        cnpj = row[P_CNPJ - 1]
        if not cnpj:
            continue
        prior[str(cnpj).strip()] = (
            row[P_SDR - 1], row[P_CANAL - 1], row[P_DETALHE - 1], row[P_STATUS - 1]
        )
    wb.close()
    return prior


def main():
    print("=" * 60)
    print("PADRONIZAÇÃO — Planilha_Nova_030726 -> layout CRM (+ Site/LinkedIn/SSMA)")
    print("=" * 60)

    prior = load_prior_control()
    if prior:
        n_imp = sum(1 for v in prior.values() if v[3] == STATUS_IMPORTED)
        print(f"Controle anterior carregado: {len(prior)} CNPJs ({n_imp} já 'Importado' preservados).")

    crm_cnpjs = load_crm_cnpjs()
    print(f"CNPJs já cadastrados no CRM: {len(crm_cnpjs)}")

    # 1) Copia as 3 linhas de cabeçalho do template padrão + coluna nova
    wb_std = openpyxl.load_workbook(TEMPLATE_STD, read_only=True)
    ws_std = wb_std[STD_SHEET]
    header_rows = []
    for r in range(1, STD_HEADER + 1):
        header_rows.append([ws_std.cell(r, c).value for c in range(1, STD_COLS + 1)])
    wb_std.close()

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = STD_SHEET
    for r, rowvals in enumerate(header_rows, start=1):
        for c, v in enumerate(rowvals, start=1):
            ws_out.cell(r, c).value = v
    # Cabeçalho das colunas novas
    ws_out.cell(2, P_LINKEDIN_EMP).value = "DADOS EXTRAS DO CLIENTE"
    ws_out.cell(STD_HEADER, P_LINKEDIN_EMP).value = "Linkedin_Empresa"
    ws_out.cell(STD_HEADER, P_NOTES_CLI).value = "Notes_Cliente"
    ws_out.cell(STD_HEADER, P_VENDEDOR).value = "Vendedor_Responsavel"

    # 2) Lê a planilha nova (NÃO read_only, para acessar hyperlinks)
    wb_src = openpyxl.load_workbook(SRC_FILE)
    ws_src = wb_src[SRC_SHEET]

    records = []
    total = com_func = com_fatur = com_site = com_linkedin = com_ssma = preservados = dup_crm = 0
    for orig, r in enumerate(range(SRC_DATA, ws_src.max_row + 1)):
        cnpj = clean(ws_src.cell(r, N_CNPJ).value)
        razao = clean(ws_src.cell(r, N_RAZAO).value)
        if not cnpj and not razao:
            continue

        fx_func  = faixa_funcionarios(ws_src.cell(r, N_FUNC).value)
        fx_fatur = faixa_faturamento(ws_src.cell(r, N_FATUR).value)
        site     = site_url(ws_src, r)
        linkedin = linkedin_empresa(ws_src, r)
        ssma     = ssma_obs(ws_src, r)
        ctrl     = prior.get(cnpj)

        com_func     += 1 if fx_func else 0
        com_fatur    += 1 if fx_fatur else 0
        com_site     += 1 if site else 0
        com_linkedin += 1 if linkedin else 0
        com_ssma      += 1 if ssma else 0
        if ctrl and ctrl[3] == STATUS_IMPORTED:
            preservados += 1
        total += 1

        records.append({
            "orig": orig, "cnpj": cnpj, "razao": razao,
            "fantasia": clean(ws_src.cell(r, N_FANTASIA).value),
            "site": site, "uf": clean(ws_src.cell(r, N_UF).value),
            "cidade": clean(ws_src.cell(r, N_CIDADE).value),
            "cnae": clean(ws_src.cell(r, N_CNAE).value),
            "fx_func": fx_func, "fx_fatur": fx_fatur,
            "fone1": clean(ws_src.cell(r, N_TEL).value),
            "email1": clean(ws_src.cell(r, N_EMAIL1).value),
            "email2": clean(ws_src.cell(r, N_EMAIL2).value),
            "ssma": ssma, "linkedin": linkedin, "ctrl": ctrl,
            "in_crm": cnpj14(cnpj) in crm_cnpjs,
        })

    wb_src.close()

    # ORDENAÇÃO das linhas (define quais os próximos lotes pegam primeiro):
    #   0) bloco já 'Importado' (ordem original) — pulado pelos lotes
    #   1) PENDENTES novos (CNPJ não está no CRM): com site primeiro, depois sem site
    #   2) PENDENTES cujo CNPJ JÁ está no CRM — jogados para o FINAL e marcados
    #      'CNPJ_no_CRM' (não sobem; ficam para revisão futura)
    def sort_key(rec):
        imported = rec["ctrl"] and rec["ctrl"][3] == STATUS_IMPORTED
        if imported:
            return (0, 0, rec["orig"])
        if rec["in_crm"]:
            return (2, 0, rec["orig"])
        return (1, 0 if rec["site"] else 1, rec["orig"])

    records.sort(key=sort_key)

    dest = STD_HEADER + 1  # começa na row 4
    for rec in records:
        ws_out.cell(dest, P_CNPJ).value      = rec["cnpj"]
        ws_out.cell(dest, P_RAZAO).value     = rec["razao"]
        ws_out.cell(dest, P_FANTASIA).value  = rec["fantasia"]
        ws_out.cell(dest, P_SITE).value      = rec["site"]
        ws_out.cell(dest, P_UF).value        = rec["uf"]
        ws_out.cell(dest, P_CIDADE).value    = rec["cidade"]
        ws_out.cell(dest, P_CNAE).value      = rec["cnae"]
        ws_out.cell(dest, P_FX_FUNC).value   = rec["fx_func"]
        ws_out.cell(dest, P_FX_FATUR).value  = rec["fx_fatur"]
        ws_out.cell(dest, P_FONE1).value     = rec["fone1"]
        ws_out.cell(dest, P_EMAIL1).value    = rec["email1"]
        ws_out.cell(dest, P_EMAIL2).value    = rec["email2"]
        ws_out.cell(dest, P_LINKEDIN_EMP).value = rec["linkedin"]
        ws_out.cell(dest, P_NOTES_CLI).value    = rec["ssma"]

        ctrl = rec["ctrl"]
        imported = ctrl and ctrl[3] == STATUS_IMPORTED
        if imported:
            # linha já subida em lote anterior — preserva o controle
            ws_out.cell(dest, P_SDR).value     = ctrl[0]
            ws_out.cell(dest, P_CANAL).value   = ctrl[1]
            ws_out.cell(dest, P_DETALHE).value = ctrl[2]
            ws_out.cell(dest, P_STATUS).value  = STATUS_IMPORTED
        elif rec["in_crm"]:
            # CNPJ já cadastrado no CRM (nunca subido por nós) — marca e joga p/ o final
            ws_out.cell(dest, P_STATUS).value  = STATUS_DUP_CRM
            dup_crm += 1
        # else: pendente novo — Status vazio

        dest += 1

    wb_out.save(OUT_FILE)

    pend_novos = total - preservados - dup_crm
    print(f"Linhas padronizadas : {total}")
    print(f"  Faixa Funcionários : {com_func}")
    print(f"  Faixa Faturamento  : {com_fatur}")
    print(f"  Site (URL)         : {com_site}")
    print(f"  LinkedIn empresa   : {com_linkedin}")
    print(f"  SSMA (notes cliente): {com_ssma}")
    print(f"  'Importado' preservados      : {preservados}")
    print(f"  'CNPJ_no_CRM' (dupes ao final): {dup_crm}")
    print(f"  Pendentes NOVOS (para subir) : {pend_novos}")
    print(f"Arquivo gerado: {os.path.basename(OUT_FILE)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
