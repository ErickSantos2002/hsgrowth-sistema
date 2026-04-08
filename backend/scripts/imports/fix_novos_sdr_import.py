"""
Script 2 de 2 — Distribuição dos primeiros 200 cards entre os novos SDRs

O que faz:
  - Lê Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx (já normalizada pelo Script 1)
  - Pega os primeiros 200 registros sem SDR preenchido
  - Distribui 50 para cada SDR: Ãhwaryoné, Miguel, Lucas, Sérgio
  - Preenche Canal_Aquisicao = "Outbound"
  - Preenche Canal_Aquisicao_Detalhe = "Outbound - Lista fria"
  - Marca as 200 linhas como "Importado" na coluna Status_Importacao
    (coluna criada ao final da planilha — serve como controle para próximas importações)
  - Salva o resultado em Planilha_Importacao_CRM_Novos_SDR_lote1.xlsx
    (arquivo que será passado para o import_from_planilha.py)
  - Também atualiza o arquivo _fixed.xlsx marcando as linhas já processadas

Rode DEPOIS do fix_novos_sdr_data.py
"""

import os
import openpyxl
from unicodedata import normalize as unicode_normalize

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
FIXED_FILE   = os.path.join(SCRIPT_DIR, "Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx")
OUTPUT_LOTE  = os.path.join(SCRIPT_DIR, "Planilha_Importacao_CRM_Novos_SDR_lote1.xlsx")

HEADER_ROW = 3
DATA_START  = 4

CARDS_PER_SDR = 50
SDRS = [
    "Ãhwaryoné",
    "Miguel",
    "Lucas",
    "Sérgio",
]

CANAL           = "Outbound"
CANAL_DETALHE   = "Outbound - Lista fria"
STATUS_IMPORTADO = "Importado"


def strip_accents(text: str) -> str:
    return unicode_normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def find_col(ws, header_row: int, name: str) -> int | None:
    name_norm = strip_accents(name.lower().strip())
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col).value
        if cell_val and strip_accents(str(cell_val).lower().strip()) == name_norm:
            return col
    return None


def main():
    print("=" * 60)
    print("SCRIPT 2 — Distribuição SDR + Canal + Status")
    print("=" * 60)
    print(f"Entrada : {FIXED_FILE}")
    print(f"Lote 1  : {OUTPUT_LOTE}")

    if not os.path.exists(FIXED_FILE):
        print("\nERRO: Arquivo _fixed.xlsx não encontrado. Rode fix_novos_sdr_data.py primeiro.")
        return

    wb = openpyxl.load_workbook(FIXED_FILE)
    ws = wb["Importação_CRM"]

    # Localiza colunas necessárias
    col_sdr    = find_col(ws, HEADER_ROW, "SDR_Responsavel *")
    col_canal  = find_col(ws, HEADER_ROW, "Canal_Aquisicao")
    col_detalhe = find_col(ws, HEADER_ROW, "Canal_Aquisicao_Detalhe")

    if not col_sdr:
        print("ERRO: Coluna 'SDR_Responsavel *' não encontrada.")
        return

    # Coluna de status: usa existente ou cria no final
    col_status = find_col(ws, HEADER_ROW, "Status_Importacao")
    if not col_status:
        col_status = ws.max_column + 1
        ws.cell(HEADER_ROW, col_status).value = "Status_Importacao"

    print(f"\nColuna SDR_Responsavel     : col {col_sdr}")
    print(f"Coluna Canal_Aquisicao     : col {col_canal}")
    print(f"Coluna Canal_Aquisicao_Det : col {col_detalhe}")
    print(f"Coluna Status_Importacao   : col {col_status}")

    # Coleta linhas ainda não importadas (Status vazio) que tenham CNPJ
    pending_rows = []
    for row in range(DATA_START, ws.max_row + 1):
        cnpj = ws.cell(row, 1).value
        if not cnpj:
            continue
        status = ws.cell(row, col_status).value
        if status == STATUS_IMPORTADO:
            continue
        pending_rows.append(row)

    total_pending = len(pending_rows)
    total_to_process = CARDS_PER_SDR * len(SDRS)  # 200

    print(f"\nLinhas pendentes (sem importar): {total_pending}")
    print(f"Linhas a processar neste lote  : {total_to_process}")

    if total_pending < total_to_process:
        print(f"\nAVISO: Apenas {total_pending} linhas disponíveis (precisaria de {total_to_process}).")
        total_to_process = total_pending

    # Distribui SDRs e preenche colunas
    rows_to_process = pending_rows[:total_to_process]
    sdr_counts = {sdr: 0 for sdr in SDRS}

    print("\nDistribuindo...")
    for i, row in enumerate(rows_to_process):
        sdr_name = SDRS[i // CARDS_PER_SDR]

        ws.cell(row, col_sdr).value = sdr_name
        if col_canal:
            ws.cell(row, col_canal).value = CANAL
        if col_detalhe:
            ws.cell(row, col_detalhe).value = CANAL_DETALHE
        ws.cell(row, col_status).value = STATUS_IMPORTADO

        sdr_counts[sdr_name] += 1

    # Salva _fixed.xlsx atualizado (com marcação de importados)
    wb.save(FIXED_FILE)
    print(f"  _fixed.xlsx atualizado com marcações 'Importado'")

    # Cria lote1 apenas com as 200 linhas selecionadas
    # (copia cabeçalhos + as linhas processadas para um arquivo separado)
    wb_lote = openpyxl.Workbook()
    ws_lote = wb_lote.active
    ws_lote.title = "Importação_CRM"

    # Copia linhas de cabeçalho (1 a DATA_START-1)
    for r in range(1, DATA_START):
        for c in range(1, ws.max_column + 1):
            ws_lote.cell(r, c).value = ws.cell(r, c).value

    # Copia as 200 linhas de dados
    for dest_row, src_row in enumerate(rows_to_process, start=DATA_START):
        for c in range(1, ws.max_column + 1):
            ws_lote.cell(dest_row, c).value = ws.cell(src_row, c).value

    wb_lote.save(OUTPUT_LOTE)

    print(f"  lote1.xlsx salvo com {len(rows_to_process)} linhas de dados")

    # Relatório final
    print("\n" + "=" * 60)
    print("RELATÓRIO")
    print("=" * 60)
    for sdr, count in sdr_counts.items():
        print(f"  {sdr:<20}: {count} cards")
    print(f"  {'TOTAL':<20}: {len(rows_to_process)} cards")
    print()
    print("Próximo passo:")
    print("  cd backend")
    print("  python scripts/imports/import_from_planilha.py Planilha_Importacao_CRM_Novos_SDR_lote1.xlsx")
    print("=" * 60)


if __name__ == "__main__":
    main()
