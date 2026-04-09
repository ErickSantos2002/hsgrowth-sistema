"""
Lote 2 — 50 leads para a Claudia (SDR id=8)

O que faz:
  - Lê Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx
  - Pega as próximas 50 linhas sem Status_Importacao
  - Atribui SDR_Responsavel = "Claudia"
  - Preenche Canal_Aquisicao = "Outbound" / Detalhe = "Outbound - Lista fria"
  - Marca como "Importado" na coluna Status_Importacao
  - Salva lote2.xlsx para passar ao import_from_planilha.py
"""

import os
import openpyxl

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
FIXED_FILE  = os.path.join(SCRIPT_DIR, "Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx")
OUTPUT_LOTE = os.path.join(SCRIPT_DIR, "Planilha_Importacao_CRM_Novos_SDR_lote2.xlsx")

HEADER_ROW      = 3
DATA_START      = 4
CARDS_TO_IMPORT = 50
SDR_NAME        = "Claudia"
CANAL           = "Outbound"
CANAL_DETALHE   = "Outbound - Lista fria"
STATUS_IMPORTADO = "Importado"


def find_col(ws, header_row, name):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip() == name:
            return col
    return None


def main():
    print("=" * 60)
    print("LOTE 2 — Claudia (SDR id=8)")
    print("=" * 60)

    wb = openpyxl.load_workbook(FIXED_FILE)
    ws = wb["Importação_CRM"]

    col_sdr     = find_col(ws, HEADER_ROW, "SDR_Responsavel *")
    col_canal   = find_col(ws, HEADER_ROW, "Canal_Aquisicao")
    col_detalhe = find_col(ws, HEADER_ROW, "Canal_Aquisicao_Detalhe")
    col_status  = find_col(ws, HEADER_ROW, "Status_Importacao")

    print(f"col SDR={col_sdr} | Canal={col_canal} | Detalhe={col_detalhe} | Status={col_status}")

    # Coleta próximas linhas livres
    pending = []
    for r in range(DATA_START, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        if ws.cell(r, col_status).value == STATUS_IMPORTADO:
            continue
        pending.append(r)

    rows_to_process = pending[:CARDS_TO_IMPORT]
    print(f"Linhas pendentes: {len(pending)} | Processando: {len(rows_to_process)}")

    for r in rows_to_process:
        ws.cell(r, col_sdr).value    = SDR_NAME
        ws.cell(r, col_canal).value  = CANAL
        ws.cell(r, col_detalhe).value = CANAL_DETALHE
        ws.cell(r, col_status).value = STATUS_IMPORTADO

    wb.save(FIXED_FILE)
    print(f"_fixed.xlsx atualizado.")

    # Gera lote2.xlsx
    wb_lote = openpyxl.Workbook()
    ws_lote = wb_lote.active
    ws_lote.title = "Importação_CRM"

    for r in range(1, DATA_START):
        for c in range(1, ws.max_column + 1):
            ws_lote.cell(r, c).value = ws.cell(r, c).value

    for dest_row, src_row in enumerate(rows_to_process, start=DATA_START):
        for c in range(1, ws.max_column + 1):
            ws_lote.cell(dest_row, c).value = ws.cell(src_row, c).value

    wb_lote.save(OUTPUT_LOTE)

    print(f"\nlote2.xlsx salvo com {len(rows_to_process)} linhas.")
    print("\nPróximo passo:")
    print("  python scripts/imports/import_from_planilha.py scripts/imports/Planilha_Importacao_CRM_Novos_SDR_lote2.xlsx")
    print("=" * 60)


if __name__ == "__main__":
    main()
