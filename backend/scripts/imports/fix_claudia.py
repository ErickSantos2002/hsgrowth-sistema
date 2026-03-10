"""
Script de correção automática da planilha Planilha_Importacao_CRM_CLAUDIA_SILVA.xlsx.

Problemas identificados (presentes em todas as 50 linhas):
  1. Estado (UF): nome completo por extenso → corrige para sigla (ex: "SAO PAULO" → "SP")
  2. Faixa de Funcionários: faixas diferentes do padrão → mapeia para o valor mais próximo
  3. Faixa de Faturamento: faixas diferentes do padrão → mapeia para o valor mais próximo
  4. Canal_Aquisicao_Detalhe: contém só "Outbound" (inválido) → limpa o campo

Gera um novo arquivo _fixed.xlsx sem modificar o original.
"""
import openpyxl
import sys
import io
from copy import copy

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

INPUT_FILE  = 'Planilha_Importacao_CRM_CLAUDIA_SILVA.xlsx'
OUTPUT_FILE = 'Planilha_Importacao_CRM_CLAUDIA_SILVA_fixed.xlsx'

# ============================================================
# Mapeamentos de correção
# ============================================================

# Estado: nome completo (sem acento, uppercase) → sigla
ESTADO_MAP = {
    'ACRE': 'AC',
    'ALAGOAS': 'AL',
    'AMAPA': 'AP',
    'AMAZONAS': 'AM',
    'BAHIA': 'BA',
    'CEARA': 'CE',
    'DISTRITO FEDERAL': 'DF',
    'ESPIRITO SANTO': 'ES',
    'GOIAS': 'GO',
    'MARANHAO': 'MA',
    'MATO GROSSO DO SUL': 'MS',
    'MATO GROSSO': 'MT',
    'MINAS GERAIS': 'MG',
    'PARA': 'PA',
    'PARAIBA': 'PB',
    'PARANA': 'PR',
    'PERNAMBUCO': 'PE',
    'PIAUI': 'PI',
    'RIO DE JANEIRO': 'RJ',
    'RIO GRANDE DO NORTE': 'RN',
    'RIO GRANDE DO SUL': 'RS',
    'RONDONIA': 'RO',
    'RORAIMA': 'RR',
    'SANTA CATARINA': 'SC',
    'SAO PAULO': 'SP',
    'SERGIPE': 'SE',
    'TOCANTINS': 'TO',
}

# Faixa de Funcionários: valor da planilha (stripped, uppercase) → padrão do sistema
FUNCIONARIOS_MAP = {
    'DE 1 A 4':        'Ate 50 colaboradores',
    'DE 5 A 9':        'Ate 50 colaboradores',
    'DE 10 A 19':      'Ate 50 colaboradores',
    'DE 20 A 49':      'Ate 50 colaboradores',
    'DE 50 A 99':      '51-100 colaboradores',
    'DE 100 A 199':    '101-300 colaboradores',
    'DE 200 A 499':    '301-600 colaboradores',
    'DE 500 A 1.000':  '601-1.000 colaboradores',
    'ACIMA DE 1.000':  'Acima de 1.000 colaboradores',
    'NAO DEFINIDO':    None,   # campo será limpo
    'NÃO DEFINIDO':    None,
}

# Faixa de Faturamento: valor da planilha (stripped) → padrão do sistema
# Estratégia: mapeia para a faixa do sistema que melhor cobre o intervalo informado
FATURAMENTO_MAP = {
    'De R$ 900 mil à R$ 4.8 milhões':        'Ate R$ 10 milhoes',
    'De R$ 4.8 milhões à R$ 10 milhões':     'Ate R$ 10 milhoes',
    'De R$ 10 milhões à R$ 15 milhões':      'R$ 10-30 milhoes',
    'De R$ 15 milhões à R$ 20 milhões':      'R$ 10-30 milhoes',
    'De R$ 20 milhões à R$ 30 milhões':      'R$ 10-30 milhoes',
    'De R$ 30 milhões à R$ 50 milhões':      'R$ 30-100 milhoes',
    'De R$ 50 milhões à R$ 75 milhões':      'R$ 30-100 milhoes',
    'De R$ 75 milhões à R$ 100 milhões':     'R$ 30-100 milhoes',
    'De R$ 100 milhões à R$ 150 milhões':    'R$ 100-300 milhoes',
    'De R$ 150 milhões à R$ 200 milhões':    'R$ 100-300 milhoes',
    'De R$ 200 milhões à R$ 300 milhões':    'R$ 100-300 milhoes',
    'De R$ 300 milhões à R$ 500 milhões':    'R$ 300 milhoes - R$ 1 bilhao',
    'De R$ 500 milhões à R$ 1 bilhão':       'R$ 300 milhoes - R$ 1 bilhao',
    'Acima de R$ 1 bilhão':                  'Acima de R$ 1 bilhao',
    'Acima de R$ 2 bilhões':                 'Acima de R$ 1 bilhao',
}

# Valores válidos do sistema (para verificação pós-correção)
VALID_ESTADOS        = {'AC','AL','AM','AP','BA','CE','DF','ES','GO','MA','MG','MS','MT',
                        'PA','PB','PE','PI','PR','RJ','RN','RO','RR','RS','SC','SE','SP','TO'}
VALID_FAIXA_FUNC     = {'Ate 50 colaboradores','51-100 colaboradores','101-300 colaboradores',
                        '301-600 colaboradores','601-1.000 colaboradores','Acima de 1.000 colaboradores'}
VALID_FAIXA_FAT      = {'Ate R$ 10 milhoes','R$ 10-30 milhoes','R$ 30-100 milhoes',
                        'R$ 100-300 milhoes','R$ 300 milhoes - R$ 1 bilhao','Acima de R$ 1 bilhao'}
VALID_CANAL_DETALHE  = {
    'Inbound - Conteudo','Inbound - Trafego pago','Inbound - SEO','Inbound - Email marketing',
    'Inbound - Levantada de mao (site / WhatsApp / formulario)',
    'Outbound - Lista fria','Outbound - LinkedIn','Outbound - Cold email','Outbound - Cold call',
    'Indicacao - Cliente','Indicacao - Parceiro','Indicacao - Network pessoal',
    'Parcerias - Co-marketing','Parcerias - Integracao tecnologica','Parcerias - Revenda',
    'Eventos - Feira','Eventos - Webinar','Eventos - Workshop','Eventos - Meetup',
    'Base - Reativacao','Base - Cross-sell','Base - Up-sell',
}

# ============================================================
# Índices das colunas (1-indexed, baseados na linha 3 da planilha)
# ============================================================
COL_ESTADO          = 6   # Estado (UF)
COL_FAIXA_FUNC      = 10  # Faixa de Funcionários
COL_FAIXA_FAT       = 11  # Faixa de Faturamento
COL_CANAL_DETALHE   = 34  # Canal_Aquisicao_Detalhe


def normalize(value: str) -> str:
    """Remove espaços, converte para uppercase e remove acentos simples para comparação."""
    import unicodedata
    s = str(value).strip().upper()
    # Remove acentos para comparação robusta
    s = ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )
    return s


def fix_estado(value) -> str | None:
    """Converte nome de estado por extenso para sigla."""
    if value is None:
        return value
    v = normalize(str(value))
    # Se já é sigla válida, mantém
    if v in VALID_ESTADOS:
        return v
    return ESTADO_MAP.get(v, value)  # devolve original se não mapeado


def fix_faixa_func(value) -> str | None:
    """Mapeia faixa de funcionários para o padrão do sistema."""
    if value is None:
        return value
    v = normalize(str(value))
    if str(value).strip() in VALID_FAIXA_FUNC:
        return str(value).strip()  # já está correto
    return FUNCIONARIOS_MAP.get(v, value)


def fix_faixa_fat(value) -> str | None:
    """Mapeia faixa de faturamento para o padrão do sistema."""
    if value is None:
        return value
    v = str(value).strip()
    if v in VALID_FAIXA_FAT:
        return v  # já está correto
    return FATURAMENTO_MAP.get(v, value)


def fix_canal_detalhe(value) -> str | None:
    """
    Limpa o campo se contiver apenas "Outbound" (valor inválido para este campo).
    Valores já válidos são mantidos.
    """
    if value is None:
        return value
    v = str(value).strip()
    if v in VALID_CANAL_DETALHE:
        return v  # já está correto
    # Se for apenas o canal sem detalhe, limpa
    if normalize(v) in ('OUTBOUND', 'INBOUND', 'INDICACAO', 'PARCERIAS', 'EVENTOS', 'BASE'):
        return None
    return value  # valor desconhecido: mantém para revisão manual


# ============================================================
# Execução
# ============================================================

wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.worksheets[0]

corrections = {
    'estado': 0,
    'faixa_func': 0,
    'faixa_fat': 0,
    'canal_detalhe': 0,
}
rows_fixed = 0
rows_total = 0

for r in range(5, ws.max_row + 1):
    # Só processa linhas com col A preenchida
    if ws.cell(r, 1).value is None:
        continue

    rows_total += 1
    row_had_fix = False

    # --- Estado (UF) ---
    cell = ws.cell(r, COL_ESTADO)
    original = cell.value
    fixed = fix_estado(original)
    if fixed != original:
        cell.value = fixed
        corrections['estado'] += 1
        row_had_fix = True

    # --- Faixa de Funcionários ---
    cell = ws.cell(r, COL_FAIXA_FUNC)
    original = cell.value
    fixed = fix_faixa_func(original)
    if fixed != original:
        cell.value = fixed
        corrections['faixa_func'] += 1
        row_had_fix = True

    # --- Faixa de Faturamento ---
    cell = ws.cell(r, COL_FAIXA_FAT)
    original = cell.value
    fixed = fix_faixa_fat(original)
    if fixed != original:
        cell.value = fixed
        corrections['faixa_fat'] += 1
        row_had_fix = True

    # --- Canal Aquisicao Detalhe ---
    cell = ws.cell(r, COL_CANAL_DETALHE)
    original = cell.value
    fixed = fix_canal_detalhe(original)
    if fixed != original:
        cell.value = fixed
        corrections['canal_detalhe'] += 1
        row_had_fix = True

    if row_had_fix:
        rows_fixed += 1

# Salva o arquivo corrigido
wb.save(OUTPUT_FILE)

# ============================================================
# Relatório
# ============================================================
sep = '=' * 60
print(sep)
print('CORRECAO AUTOMATICA - PLANILHA CLAUDIA SILVA')
print(sep)
print(f'Arquivo gerado : {OUTPUT_FILE}')
print(f'Linhas totais  : {rows_total}')
print(f'Linhas corrigidas: {rows_fixed}')
print()
print('Correcoes por campo:')
print(f'  Estado (UF)              : {corrections["estado"]} celulas')
print(f'  Faixa de Funcionarios    : {corrections["faixa_func"]} celulas')
print(f'  Faixa de Faturamento     : {corrections["faixa_fat"]} celulas')
print(f'  Canal_Aquisicao_Detalhe  : {corrections["canal_detalhe"]} celulas (campo limpo)')
print()
print('AVISO: Canal_Aquisicao_Detalhe foi LIMPO nas linhas onde continha')
print('apenas "Outbound" (valor invalido para esse campo).')
print('Se houver detalhe especifico (Cold call, LinkedIn, etc.),')
print('preencha manualmente antes de importar.')
print(sep)
