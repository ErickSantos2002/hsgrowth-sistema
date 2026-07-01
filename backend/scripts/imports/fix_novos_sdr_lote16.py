"""
Lote 16 — 159 leads (esgota a planilha): dividido entre Karolaine e Miguel
  - Karolaine Martins                    (id=9)   → 80 cards
  - Miguel Luiz Pereira de Melo         (id=16)  → 79 cards

Divisão ímpar de 159: Karolaine leva 1 a mais.

O que faz:
  - Lê Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx
  - Pega TODAS as linhas restantes sem Status_Importacao (até 159)
  - Distribui conforme SDR_COUNTS
  - Preenche Canal_Aquisicao = "Outbound" / Detalhe = "Outbound - Lista fria"
  - Marca como "Importado" na coluna Status_Importacao
  - Salva lote16.xlsx para passar ao import_from_planilha.py
"""

import os
import openpyxl

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
FIXED_FILE  = os.path.join(SCRIPT_DIR, "Planilha_Importacao_CRM_Novos_SDR_fixed.xlsx")
OUTPUT_LOTE = os.path.join(SCRIPT_DIR, "Planilha_Importacao_CRM_Novos_SDR_lote16.xlsx")

HEADER_ROW       = 3
DATA_START       = 4
CANAL            = "Outbound"
CANAL_DETALHE    = "Outbound - Lista fria"
STATUS_IMPORTADO = "Importado"

# (nome, quantidade) — metade para cada, Karolaine leva o card extra do total ímpar
SDR_COUNTS = [
    ("Karolaine", 80),
    ("Miguel",    79),
]


def find_col(ws, header_row, name):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip() == name:
            return col
    return None


def main():
    print("=" * 60)
    print("LOTE 16 — 159 cards (Karolaine 80 + Miguel 79) — esgota a planilha")
    print("=" * 60)

    wb = openpyxl.load_workbook(FIXED_FILE)
    ws = wb["Importação_CRM"]

    col_sdr     = find_col(ws, HEADER_ROW, "SDR_Responsavel *")
    col_canal   = find_col(ws, HEADER_ROW, "Canal_Aquisicao")
    col_detalhe = find_col(ws, HEADER_ROW, "Canal_Aquisicao_Detalhe")
    col_status  = find_col(ws, HEADER_ROW, "Status_Importacao")

    print(f"col SDR={col_sdr} | Canal={col_canal} | Detalhe={col_detalhe} | Status={col_status}")

    # Coleta próximas linhas livres
    pending = []
    for r in range(DATA_START, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        if ws.cell(r, col_status).value == STATUS_IMPORTADO:
            continue
        pending.append(r)

    total_needed = sum(n for _, n in SDR_COUNTS)
    print(f"Linhas pendentes: {len(pending)} | Necessárias: {total_needed}")

    # Constrói a sequência de SDRs conforme as quantidades
    assignment = []
    for name, n in SDR_COUNTS:
        assignment.extend([name] * n)

    # Se houver menos linhas do que o esperado, ajusta (não deve acontecer)
    rows_to_process = pending[:total_needed]
    assignment = assignment[:len(rows_to_process)]

    for r, sdr_name in zip(rows_to_process, assignment):
        ws.cell(r, col_sdr).value     = sdr_name
        ws.cell(r, col_canal).value   = CANAL
        ws.cell(r, col_detalhe).value = CANAL_DETALHE
        ws.cell(r, col_status).value  = STATUS_IMPORTADO

    wb.save(FIXED_FILE)
    print(f"_fixed.xlsx atualizado.")

    # Resumo por SDR
    print("\nDistribuição:")
    for name, _ in SDR_COUNTS:
        print(f"  {name}: {assignment.count(name)} cards")

    # Gera lote16.xlsx
    wb_lote = openpyxl.Workbook()
    ws_lote = wb_lote.active
    ws_lote.title = "Importação_CRM"

    for r in range(1, DATA_START):
        for c in range(1, ws.max_column + 1):
            ws_lote.cell(r, c).value = ws.cell(r, c).value

    # Row 4 fica em branco intencionalmente — import_from_planilha.py pula a primeira linha de dados (row 4),
    # então os dados começam na row 5 para que todos os cards sejam importados sem perda.
    for dest_row, src_row in enumerate(rows_to_process, start=DATA_START + 1):
        for c in range(1, ws.max_column + 1):
            ws_lote.cell(dest_row, c).value = ws.cell(src_row, c).value

    wb_lote.save(OUTPUT_LOTE)

    print(f"\nlote16.xlsx salvo com {len(rows_to_process)} linhas (dados a partir da row 5 — sem card faltante).")
    print("\nPróximo passo:")
    print("  python scripts/imports/import_from_planilha.py C:/Users/HS/Documents/GitHub/hsgrowth-sistema/backend/scripts/imports/Planilha_Importacao_CRM_Novos_SDR_lote16.xlsx")
    print("=" * 60)


if __name__ == "__main__":
    main()
