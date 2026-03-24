"""
fix_cnpj_karolaine.py — Corrige os CNPJs dos clientes importados via Karolaine.

Problema:
  Na planilha original, a coluna CNPJ estava deslocada 1 linha para baixo em
  relação à Razão Social. O script de importação leu ambas da mesma linha, então:
    - Cliente da linha N recebeu o CNPJ que pertencia ao cliente da linha N-1
    - SERRA VERDE (linha 5) ficou sem CNPJ
    - OPENEEM (linha 54) recebeu CNPJ errado (o correto estava na linha 55)

Correção:
  Para cada empresa na linha N, o CNPJ correto é o da linha N+1 no xlsx.
  O script encontra cada cliente no banco pelo nome exato e atualiza o campo
  document com o CNPJ correto (14 dígitos, zero-padded).

Uso:
  cd backend
  python scripts/imports/fix_cnpj_karolaine.py [--dry-run]
"""

import os
import sys
import argparse

import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

from app.db.session import SessionLocal
from app.models.client import Client


def clean_cnpj(val) -> str | None:
    if val is None:
        return None
    digits = "".join(c for c in str(val).strip() if c.isdigit())
    return digits.zfill(14) if digits else None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true",
                        help="Apenas mostra o que seria feito, sem salvar no banco")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(script_dir, "Planilha_Importacao_CRM_Karolaine_fixed.xlsx")

    if not os.path.exists(xlsx_path):
        print(f"ERRO: arquivo não encontrado: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Localiza colunas de Razão Social e CNPJ (linha 3 = cabeçalho)
    col_rs = col_cnpj = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=3, column=c).value
        if v == "Razão Social *":
            col_rs = c
        if v == "CNPJ *":
            col_cnpj = c

    if not col_rs or not col_cnpj:
        print("ERRO: colunas 'Razão Social *' ou 'CNPJ *' não encontradas.")
        sys.exit(1)

    # Monta lista de (razao_social, cnpj_correto) onde cnpj_correto vem da linha N+1
    # Dados começam na linha 5; última linha com dados é 54 (linha 55 tem RS vazia)
    corrections = []
    for r in range(5, ws.max_row + 1):
        razao = ws.cell(row=r, column=col_rs).value
        if not razao or not str(razao).strip():
            continue
        razao = str(razao).strip()
        # CNPJ correto está na próxima linha
        cnpj_raw = ws.cell(row=r + 1, column=col_cnpj).value
        cnpj_correct = clean_cnpj(cnpj_raw)
        corrections.append((razao, cnpj_correct))

    print("=" * 70)
    print("FIX CNPJ KAROLAINE")
    print("=" * 70)
    print(f"{'Razão Social':<50} {'CNPJ Correto'}")
    print("-" * 70)
    for rs, cnpj in corrections:
        print(f"{rs[:49]:<50} {cnpj or '(sem CNPJ)'}")
    print(f"\nTotal de empresas a corrigir: {len(corrections)}")

    if args.dry_run:
        print("\n[DRY RUN] Nenhuma alteração foi feita no banco.")
        return

    db = SessionLocal()
    try:
        ok = 0
        not_found = []
        for razao, cnpj_correct in corrections:
            client = (
                db.query(Client)
                .filter(Client.name == razao, Client.is_deleted == False)
                .first()
            )
            if not client:
                not_found.append(razao)
                continue
            old_doc = client.document or "(vazio)"
            client.document = cnpj_correct
            ok += 1
            print(f"  [OK] {razao[:45]:<46} {old_doc} -> {cnpj_correct or '(vazio)'}")

        db.commit()
        print("\n" + "=" * 70)
        print(f"  Clientes corrigidos : {ok}")
        if not_found:
            print(f"  Não encontrados ({len(not_found)}):")
            for n in not_found:
                print(f"    - {n}")
        print("=" * 70)
    except Exception as e:
        db.rollback()
        print(f"\nERRO: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
